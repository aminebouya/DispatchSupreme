import os
import pandas as pd
import numpy as np
import streamlit as st
import time
from geopy.geocoders import GoogleV3
from geopy.exc import GeocoderTimedOut, GeocoderQuotaExceeded
from geopy.distance import geodesic
import folium
from folium.features import DivIcon
import streamlit.components.v1 as components
import warnings
import urllib.parse
import io
import googlemaps
from polyline import decode
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.filters import FilterColumn, Filters
import html
import re
from datetime import datetime

# Suppress openpyxl warning about default style
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")

# ─── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="Dispatch Supreme", layout="wide")

# ─── GLOBAL STYLES ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
  /* Reduce overall Streamlit page size to 90% */
  .main .block-container {
      max-width: 90% !important;
      padding-left: 1rem !important;
      padding-right: 1rem !important;
  }

  /* Decrease main title size slightly, still larger than subheader */
  .main .block-container h1 {
      font-size: 18px !important;
  }

  /* Make "FILTERS" heading same size and style as table titles */
  h2.filters-heading {
      font-size: 22px !important;
      font-weight: bold !important;
      text-decoration: underline;
  }

  .filters-label {
    font-size: 16px !important;
    font-weight: bold !important;
    margin-bottom: -20px !important;
    display: block;
    line-height: 1;
  }

  .name-filters-label {
    font-size: 14px !important;
    font-weight: bold !important;
    margin-bottom: -20px !important;
    display: block;
    line-height: 1;
  }

  /* Ensure button text is bold */
  .stButton>button {
      font-weight: bold !important;
  }

  [data-testid="stSlider"] .rc-slider { width: 200px !important; }
  [data-testid="stSlider"] .rc-slider-handle {
      font-size: 12px !important;
      font-weight: bold !important;
  }
  /* Table header styling */
  .stDataFrame table {
      border-collapse: collapse;
      width: 100%;
      font-size: 12px;
  }
  .stDataFrame th {
      background-color: #e6f3ff;
      color: #003087;
      font-weight: bold;
      padding: 10px;
      font-size: 14px !important;
  }
  .stDataFrame td {
      padding: 8px;
      border: 1px solid #ddd;
      font-size: 10px !important;
  }
  .stDataFrame tr:nth-child(even) {
      background-color: #f0f8ff;
  }
  .stDataFrame tr:hover {
      background-color: #d9e6ff;
  }
  
  /* Center content in Drivers and Priority columns in data editor */
  div[data-testid="stDataFrameResizable"] td:nth-child(1),
  div[data-testid="stDataFrameResizable"] td:nth-child(3) {
      text-align: center !important;
  }

  /* Target Streamlit widget containers for minimal vertical spacing */
  div[data-testid="stSelectbox"],
  div[data-testid="stMultiSelect"],
  div[data-testid="stSlider"],
  div[data-testid="stRadio"],
  div[data-testid="stCheckbox"],
  .stCheckbox,
  .stRadio,
  div[data-testid="stTextInput"] {
    margin-top: -20px !important;
    margin-bottom: 0px !important;
    padding-top: 0px !important;
    padding-bottom: 0px !important;
  }

  /* Specific styling for the input field inside the warehouse selectbox */
  [data-testid="stSelectbox"] input {
      width: 5ch !important;
  }

  /* Control overall width of the selectbox widget for Whs */
  div[data-testid="stVerticalBlock"] > div > div > div[data-testid="stSelectbox"] {
      width: auto !important;
      min-width: unset !important;
      max-width: 100% !important;
  }

  /* Specific styling for driver name input field */
  div[data-testid="stTextInput"][data-key^="driver_name_"] input {
      width: 10ch !important;
  }

  /* Specific styling for first driver name input to add spacing */
  div[data-testid="stTextInput"][data-key="driver_name_0"] {
      margin-top: 10px !important;
  }

  /* Ensure checkbox labels do not wrap */
  div[data-testid="stCheckbox"] label span {
      white-space: nowrap;
  }

  /* Custom styles for bolding upload file labels */
  .stFileUploader label {
    font-weight: bold !important;
    font-size: 16px !important;
  }

  /* Style for zoom controls */
  .leaflet-control-zoom {
      background-color: white !important;
      border: 2px solid #333 !important;
      border-radius: 4px !important;
      padding: 5px !important;
      position: absolute !important;
      top: 10px !important;
      left: 10px !important;
      z-index: 1000 !important;
  }
  .leaflet-control-zoom a {
      font-size: 20px !important;
      font-weight: bold !important;
      color: #333 !important;
      text-decoration: none !important;
      display: block !important;
      text-align: center !important;
      width: 30px !important;
      height: 30px !important;
      line-height: 30px !important;
      margin: 2px 0 !important;
  }
  .leaflet-control-zoom a:hover {
      background-color: #f0f0f0 !important;
  }

  /* Ensure radio button text is bold */
  [data-testid="stRadio"] label {
      font-weight: bold !important;
  }
  [data-testid="stRadio"] label span {
      font-weight: bold !important;
  }

  /* Style for expanders */
  .streamlit-expanderHeader {
      font-size: 19px !important;
      font-weight: bold !important;
      color: #003087 !important;
      background-color: #f0f8ff !important;
      padding: 10px !important;
  }

  /* Style for Planned Driving Trajectories label */
  .trajectories-label {
    font-size: 18px !important;
    font-weight: bold !important;
    margin-bottom: -20px !important;
    display: block;
    line-height: 1;
    padding-bottom: 10px;
  }

  /* Global Button Styling - Colorful buttons for all except Trajectory Visibility */
  div[data-testid="stButton"] > button:not([key*="show_all"]):not([key*="hide_all"]) {
    background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%) !important;
    color: white !important;
    font-weight: bold !important;
    border-radius: 6px !important;
    border: none !important;
    box-shadow: 0 3px 6px rgba(0,0,0,0.15) !important;
    transition: all 0.2s ease !important;
  }
  div[data-testid="stButton"] > button:not([key*="show_all"]):not([key*="hide_all"]):hover {
    background: linear-gradient(135deg, #45a049 0%, #3d8b40 100%) !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.2) !important;
  }

  /* Special styling for primary buttons (like Add All Trajectories) */
  div[data-testid="stButton"] > button[kind="primary"] {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
    color: white !important;
    font-weight: bold !important;
    border-radius: 8px !important;
    border: none !important;
    height: 3em !important;
    font-size: 1.1em !important;
    box-shadow: 0 4px 8px rgba(0,0,0,0.2) !important;
    transition: all 0.3s ease !important;
  }
  div[data-testid="stButton"] > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #5a6fd8 0%, #6a4190 100%) !important;
    transform: translateY(-2px) !important;
    box-shadow: 0 6px 12px rgba(0,0,0,0.3) !important;
  }

  /* Removed problematic grey button CSS - using JavaScript instead */

  /* Dark mode compatibility for grey button */
  @media (prefers-color-scheme: dark) {
    div[data-testid="stButton"] > button[key="clear_pasted_trajectory"] {
      background: linear-gradient(135deg, #495057 0%, #343a40 100%) !important;
      color: #f8f9fa !important;
      border: 1px solid #6c757d !important;
    }
    div[data-testid="stButton"] > button[key="clear_pasted_trajectory"]:hover {
      background: linear-gradient(135deg, #343a40 0%, #212529 100%) !important;
      border: 1px solid #adb5bd !important;
    }
  }
</style>
""", unsafe_allow_html=True)

# --- Warehouse Coordinates ---
WAREHOUSE_COORDS = {
    "C01": {"lat": 41.721109, "lon": -81.266314, "name": "C01 Warehouse"},
    "D01": {"lat": 42.361272, "lon": -83.186465, "name": "D01 Warehouse"},
    "S01": {"lat": 39.299065, "lon": -84.431747, "name": "S01 Warehouse"},
    "P01": {"lat": 40.485879, "lon": -80.073695, "name": "P01 Warehouse"}
}

# --- Function to sanitize text for HTML ---
def sanitize_for_html(text):
    if pd.isna(text):
        return ""
    # Replace problematic characters and escape HTML
    text = str(text).replace("`", "'")  # Replace ` with '
    text = re.sub(r'[^\x00-\x7F]+', '', text)  # Remove non-ASCII characters
    return html.escape(text)

# --- Function to sanitize sheet name for Excel ---
def sanitize_sheet_name(name):
    if not name:
        return "Sheet1"
    # Remove invalid characters: /, \, *, ?, :, [, ]
    invalid_chars = r'[\/*?:[\]]'
    sanitized = re.sub(invalid_chars, '_', str(name))
    # Truncate to 31 characters (Excel limit)
    sanitized = sanitized[:31]
    # Ensure not empty after sanitization
    return sanitized or "Sheet1"

# --- Function to parse Google Maps URL ---
def parse_google_maps_url(url):
    try:
        parsed_url = urllib.parse.urlparse(url)
        query_params = urllib.parse.parse_qs(parsed_url.query)
        
        origin = query_params.get('origin', [''])[0]
        destination = query_params.get('destination', [''])[0]
        waypoints = query_params.get('waypoints', [''])[0].split('|') if query_params.get('waypoints') else []
        
        origin = urllib.parse.unquote(origin).strip()
        destination = urllib.parse.unquote(destination).strip()
        waypoints = [urllib.parse.unquote(wp).strip() for wp in waypoints if wp.strip()]
        
        addresses = [origin] + waypoints + [destination]
        addresses = [addr for addr in addresses if addr]
        
        return addresses
    except Exception as e:
        st.warning(f"Error parsing Google Maps URL: {e}")
        return []

# --- Function to geocode addresses ---
def geocode_addresses(addresses, cache, geolocator):
    coords = []
    updated_cache = cache.copy()
    invalid_addresses = []
    
    for addr in addresses:
        if addr in updated_cache.index:
            lat, lon = updated_cache.loc[addr, ['lat', 'lon']]
            if not pd.isna(lat) and not pd.isna(lon):
                coords.append((lat, lon))
                continue
        
        try:
            loc = geolocator.geocode(addr, timeout=10)
            if loc:
                lat, lon = loc.latitude, loc.longitude
                coords.append((lat, lon))
                updated_cache.loc[addr] = {'lat': lat, 'lon': lon, 'Name': ''}
            else:
                invalid_addresses.append(addr)
        except (GeocoderTimedOut, GeocoderQuotaExceeded) as e:
            invalid_addresses.append(addr)
    
    if not updated_cache.empty:
        updated_cache.to_csv(CACHE_PATH, encoding='utf-8')
    return coords, updated_cache, invalid_addresses

# --- Function to format travel time for display ---
def format_travel_time(minutes):
    """
    Convert minutes to human-readable format (e.g., "2h 30m")
    """
    if minutes == 0:
        return "0m"
    
    hours = int(minutes // 60)
    mins = int(minutes % 60)
    
    if hours == 0:
        return f"{mins}m"
    elif mins == 0:
        return f"{hours}h"
    else:
        return f"{hours}h {mins}m"

# --- Function to optimize delivery order for truck routing ---
def optimize_delivery_order(driver_data, warehouse_coords):
    """
    Optimize delivery order using nearest neighbor approach with geodesic distance
    to minimize travel distance for truck routes
    """
    if len(driver_data) <= 1:
        return driver_data
    
    try:
        from geopy.distance import geodesic
        
        # Convert to list for easier manipulation
        deliveries = driver_data.copy().reset_index(drop=True)
        optimized_order = []
        remaining = list(range(len(deliveries)))
        
        # Start from the delivery closest to warehouse
        current_pos = warehouse_coords
        
        while remaining:
            # Find nearest unvisited delivery using geodesic distance
            min_distance = float('inf')
            nearest_idx = None
            nearest_pos = None
            
            for idx in remaining:
                delivery = deliveries.iloc[idx]
                if pd.notna(delivery["lat_jitter"]) and pd.notna(delivery["lon_jitter"]):
                    delivery_pos = [delivery["lat_jitter"], delivery["lon_jitter"]]
                    # Use geodesic distance for more accurate calculation
                    distance = geodesic(current_pos, delivery_pos).miles
                    
                    if distance < min_distance:
                        min_distance = distance
                        nearest_idx = idx
                        nearest_pos = delivery_pos
            
            if nearest_idx is not None:
                optimized_order.append(nearest_idx)
                remaining.remove(nearest_idx)
                current_pos = nearest_pos
        
        # Return deliveries in optimized order
        optimized_result = deliveries.iloc[optimized_order].reset_index(drop=True)
        return optimized_result
        
    except Exception as e:
        # Return original order if optimization fails
        return driver_data

# --- Function to get driving route ---
def get_driving_route(addresses, gmaps_client):
    if len(addresses) < 2:
        return []
    
    try:
        waypoints = addresses[1:-1] if len(addresses) > 2 else []
        directions_result = gmaps_client.directions(
            origin=addresses[0],
            destination=addresses[-1],
            waypoints=waypoints,
            mode="driving",
            optimize_waypoints=False  # Keep our priority-based order
        )
        
        if not directions_result:
            return []
        
        polyline = directions_result[0]['overview_polyline']['points']
        route_coords = decode(polyline)
        return route_coords
    except Exception as e:
        return []

# --- Function to load geocode cache ---
CACHE_PATH = "geocode_cache.csv"
def load_cache():
    if os.path.exists(CACHE_PATH):
        encodings = ['utf-8', 'latin1', 'windows-1252']
        for encoding in encodings:
            try:
                cache = pd.read_csv(CACHE_PATH, index_col="full_address", encoding=encoding)
                if 'Name' not in cache.columns:
                    cache['Name'] = ''
                if encoding != 'utf-8':
                    st.info(f"Successfully loaded cache with {encoding} encoding.")
                return cache
            except (UnicodeDecodeError, pd.errors.ParserError) as e:
                st.warning(f"Failed to load cache with {encoding} encoding: {e}")
        st.error(f"Could not read {CACHE_PATH} with any supported encoding. Creating a new empty cache.")
        return pd.DataFrame(columns=["full_address", "lat", "lon", "Name"]).set_index("full_address")
    else:
        return pd.DataFrame(columns=["full_address", "lat", "lon", "Name"]).set_index("full_address")

# --- Function to transform email usernames ---
def transform_username(username):
    username = username.strip().lower()
    transformations = {
        "jbartnic": "jbartnicki",
        "cdaugher": "cdaugherty",
        "bwilderm": "bwildermuth",
        "jhigginb": "jhigginbotham",
        "rmothers": "rmothersell",
        "sferguso": "sferguson"
    }
    return transformations.get(username, username)

# --- Function to find deliveries near the trajectory ---
def find_nearby_deliveries(df_geo, trajectory_coords, threshold_miles, show_real_only=True, map_date_filter=[]):
    if not trajectory_coords or len(trajectory_coords) < 2:
        return pd.DataFrame()
    
    if map_date_filter:
        df_geo = df_geo[df_geo["Date"].isin(map_date_filter)].copy()
    
    nearby_deliveries = []
    inventory_df = st.session_state.inventory_df.copy()
    
    for idx, row in df_geo.iterrows():
        delivery_coords = (row["lat_jitter"], row["lon_jitter"])
        min_distance = min(
            geodesic(delivery_coords, traj_point).miles
            for traj_point in trajectory_coords
        )
        if min_distance <= threshold_miles:
            on_hand_approve = 0
            allocated_qty = 0
            open_orders = 0
            calculated = 0

            if not inventory_df.empty:
                item_no_str = str(row["Item no"])
                whs_str = str(row["Whs"])
                inventory_match = inventory_df[
                    (inventory_df["Item number"] == item_no_str) &
                    (inventory_df["Whs"] == whs_str)
                ]
                if not inventory_match.empty:
                    on_hand_approve_val = inventory_match["On-hand approve"].iloc[0]
                    allocated_qty_val = inventory_match["Allocated qty"].iloc[0] if "Allocated qty" in inventory_match.columns else 0
                    try:
                        on_hand_approve = float(on_hand_approve_val) if pd.notna(on_hand_approve_val) and str(on_hand_approve_val).strip() != '' else 0
                        allocated_qty = float(allocated_qty_val) if pd.notna(allocated_qty_val) and str(allocated_qty_val).strip() != '' else 0
                        earlier_orders = [r["Qty"] for r in df_geo.to_dict(orient="records") if r["Item no"] == item_no_str and r["Date"] < row["Date"]]
                        open_orders = sum(earlier_orders) if earlier_orders else 0
                        calculated = on_hand_approve - (allocated_qty + open_orders + row["Qty"])
                    except ValueError:
                        pass
            
            nearby_deliveries.append({
                "Date": row["Date"],
                "Resp": row.get("Resp", ""),
                "Salespers": row.get("Salespers", ""),
                "CO no": row["CO no"],
                "Name": row["Name"],
                "Full Address": row["full_address"],
                "Item Number": row["Item no"],
                "Product": row["Product"],
                "Unit": row["U/M"],
                "Quantity": row["Qty"],
                "Gross Weight": round(row["Qty"] * row["Gross wt"], 2),
                "Distance": round(min_distance, 2),
                "On-Hand": on_hand_approve,
                "Allocated Qty": allocated_qty,
                "Open Orders": open_orders,
                "Calculated": calculated
            })
    
    if not nearby_deliveries:
        return pd.DataFrame()
    
    nearby_df = pd.DataFrame(nearby_deliveries)
    nearby_df["Date"] = pd.to_datetime(nearby_df["Date"], format="%m-%d-%Y")
    nearby_df = nearby_df.sort_values(["Date", "Distance"])
    nearby_df["Date"] = nearby_df["Date"].dt.strftime("%m-%d-%Y")

    if show_real_only and "Calculated" in nearby_df.columns:
        nearby_df = nearby_df[
            pd.to_numeric(nearby_df["Calculated"], errors='coerce').notna() &
            (pd.to_numeric(nearby_df["Calculated"], errors='coerce') > 0)
        ]

    return nearby_df

# --- Function to generate Excel file for download ---
def generate_excel_file(df, sheet_name="Sheet1", align_left=False):
    if df.empty:
        return None
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sanitize_sheet_name(sheet_name)

    headers = df.columns.tolist()
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left' if align_left else 'center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for row_idx, row in enumerate(df.values, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left' if align_left else 'general')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    ws.auto_filter.ref = ws.dimensions
    wb.save(output)
    return output.getvalue()

# --- Function to perform geocoding and initial plan generation ---
def generate_initial_plan_data(df_to_process, inventory_df, current_cache):
    df_to_process["full_address"] = (
        df_to_process["Address 1"].fillna('').str.strip() + ", " +
        df_to_process["City"].fillna('').str.strip() + ", " +
        df_to_process["Sta"].fillna('').str.strip()
    )

    df_to_process = df_to_process[df_to_process["full_address"].str.replace(',\\s*', '', regex=True).astype(bool)]
    if df_to_process.empty:
        return pd.DataFrame(), {}, pd.DataFrame(), current_cache

    if "Geo code X" in df_to_process.columns and "Geo code Y" in df_to_process.columns:
        df_to_process["lon"] = pd.to_numeric(df_to_process["Geo code X"], errors="coerce")
        df_to_process["lat"] = pd.to_numeric(df_to_process["Geo code Y"], errors="coerce")
    else:
        df_to_process["lon"] = np.nan
        df_to_process["lat"] = np.nan

    missing_coords = df_to_process[df_to_process["lat"].isna() | df_to_process["lon"].isna()]
    if not missing_coords.empty:
        geolocator = GoogleV3(api_key=st.secrets["GOOGLE_API_KEY"], timeout=10)

        for idx, row in missing_coords.iterrows():
            addr = row["full_address"]
            if addr in current_cache.index:
                lat, lon = current_cache.loc[addr, ['lat', 'lon']]
                if pd.notna(lat) and pd.notna(lon):
                    df_to_process.at[idx, "lat"] = lat
                    df_to_process.at[idx, "lon"] = lon
                    continue
            try:
                loc = geolocator.geocode(addr, timeout=10)
                if loc:
                    lat, lon = loc.latitude, loc.longitude
                    df_to_process.at[idx, "lat"] = lat
                    df_to_process.at[idx, "lon"] = lon
                    current_cache.loc[addr] = {'lat': lat, 'lon': lon, 'Name': row["Name"]}
                else:
                    pass
            except (GeocoderTimedOut, GeocoderQuotaExceeded) as e:
                pass

        current_cache.to_csv(CACHE_PATH, encoding="utf-8")

    df_geo = df_to_process.dropna(subset=["lat", "lon"]).copy()
    df_geo["lat_jitter"] = df_geo["lat"] + np.random.uniform(-0.0005, 0.0005, len(df_geo))
    df_geo["lon_jitter"] = df_geo["lon"] + np.random.uniform(-0.0005, 0.0005, len(df_geo))

    df_geo["Product"] = df_geo["Description 1"].astype(str)
    df_geo["Unit"] = df_geo["U/M"].astype(str)
    df_geo["Qty"] = df_geo["Order qty bU/M"]
    df_geo["Item no"] = df_geo["Item no"].astype(str)
    df_geo["CO no"] = df_geo["CO no"].astype(str)
    
    # Ensure Route column exists and is properly handled
    if "Route" not in df_geo.columns:
        df_geo["Route"] = ""
    df_geo["Route"] = df_geo["Route"].fillna("").astype(str)

    dates = sorted(df_geo["Date"].unique())
    palette = [
        [255, 0, 0], [0, 128, 0], [0, 0, 255], [255, 165, 0], [128, 0, 128],
        [255, 192, 203], [0, 255, 255], [255, 255, 0], [139, 69, 19], [128, 128, 128],
        [255, 99, 71], [0, 191, 255], [75, 0, 130], [173, 255, 47], [220, 20, 60]
    ]
    # Create color mapping with black for the first date
    cmap = {}
    for i, d in enumerate(dates):
        if i == 0:
            # First date gets black color
            cmap[d] = [0, 0, 0, 200]  # Black with alpha
        else:
            # Other dates use the palette (shifted by 1 to skip the red)
            cmap[d] = palette[(i-1) % len(palette)] + [200]
    df_geo["color"] = df_geo["Date"].map(cmap)

    plan = []
    recs = df_geo.to_dict(orient="records")
    inventory_df = inventory_df.copy()

    if len(recs) < 2:
        return df_geo, cmap, pd.DataFrame(columns=[
            "B Date", "B Name", "B Address",
            "C Date", "C CO#", "C Item#", "C Product", "C U/M", "C Qty", "C Gross Wt",
            "C Name", "C Address", "Distance", "On-Hand", "Allocated Qty", "Open Orders", "Calculated",
            "C Resp", "C Salespers"
        ]), current_cache

    for base in recs:
        for cand in recs:
            if cand["Date"] > base["Date"]:
                dist = 0.0 if base["full_address"] == cand["full_address"] else geodesic(
                    (base["lat_jitter"], base["lon_jitter"]),
                    (cand["lat_jitter"], cand["lon_jitter"])
                ).miles
                on_hand_approve = 0
                allocated_qty = 0
                open_orders = 0
                calculated = 0

                if not inventory_df.empty:
                    cand_item_no_str = str(cand["Item no"])
                    cand_whs_str = str(cand["Whs"])
                    inventory_match = inventory_df[
                        (inventory_df["Item number"] == cand_item_no_str) &
                        (inventory_df["Whs"] == cand_whs_str)
                    ]
                    if not inventory_match.empty:
                        on_hand_approve_val = inventory_match["On-hand approve"].iloc[0]
                        allocated_qty_val = inventory_match["Allocated qty"].iloc[0] if "Allocated qty" in inventory_match.columns else 0
                        try:
                            on_hand_approve = float(on_hand_approve_val) if pd.notna(on_hand_approve_val) and str(on_hand_approve_val).strip() != '' else 0
                            allocated_qty = float(allocated_qty_val) if pd.notna(allocated_qty_val) and str(allocated_qty_val).strip() != '' else 0
                            earlier_orders = [r["Qty"] for r in recs if r["Item no"] == cand["Item no"] and r["Date"] < cand["Date"]]
                            open_orders = sum(earlier_orders) if earlier_orders else 0
                            calculated = on_hand_approve - (allocated_qty + open_orders + cand["Qty"])
                        except ValueError:
                            pass

                plan.append({
                    "B Date": base["Date"],
                    "B Name": base["Name"],
                    "B Address": base["full_address"],
                    "C Date": cand["Date"],
                    "C CO#": cand["CO no"],
                    "C Item#": cand["Item no"],
                    "C Product": cand["Product"],
                    "C U/M": cand["Unit"],
                    "C Qty": cand["Qty"],
                    "C Gross Wt": round(cand["Qty"] * cand["Gross wt"], 2),
                    "C Name": cand["Name"],
                    "C Address": cand["full_address"],
                    "Distance": round(dist, 2),
                    "On-Hand": on_hand_approve,
                    "Allocated Qty": allocated_qty,
                    "Open Orders": open_orders,
                    "Calculated": calculated,
                    "C Resp": cand.get("Resp", ""),
                    "C Salespers": cand.get("Salespers", "")
                })

    if not plan:
        plan_df = pd.DataFrame(columns=[
            "B Date", "B Name", "B Address",
            "C Date", "C CO#", "C Item#", "C Product", "C U/M", "C Qty", "C Gross Wt",
            "C Name", "C Address", "Distance", "On-Hand", "Allocated Qty", "Open Orders", "Calculated",
            "C Resp", "C Salespers"
        ])
    else:
        plan_df = pd.DataFrame(plan)

    return df_geo, cmap, plan_df, current_cache

# Load cache early to ensure availability
cache = load_cache()

# Initialize Google Maps client
gmaps_client = googlemaps.Client(key=st.secrets["GOOGLE_API_KEY"])

# Renamed the main title
st.title("Dispatch Supreme")

# ─── SESSION STATE ──────────────────────────────────────────────────────────────
if "run" not in st.session_state:
    st.session_state.run = False
if "inventory_df" not in st.session_state:
    st.session_state.inventory_df = pd.DataFrame()
if "df_geo" not in st.session_state:
    st.session_state.df_geo = pd.DataFrame()
if "cmap" not in st.session_state:
    st.session_state.cmap = {}
if "ap_df" not in st.session_state:
    st.session_state.ap_df = pd.DataFrame()
if "raw_ap_df" not in st.session_state:
    st.session_state.raw_ap_df = pd.DataFrame()
if "last_filtered_df_hash" not in st.session_state:
    st.session_state.last_filtered_df_hash = None
if "trajectories" not in st.session_state:
    st.session_state.trajectories = [
        {"coords": [], "driver": "", "fill_rate": "", "lbs_to_add": "", "url": "", "added": False}
        for _ in range(7)
    ]
if "overdue_containers_df" not in st.session_state:
    st.session_state.overdue_containers_df = pd.DataFrame()
if "date_filter_hash" not in st.session_state:
    st.session_state.date_filter_hash = None
if "action_plan_email_triggered" not in st.session_state:
    st.session_state.action_plan_email_triggered = False
if "overdue_email_triggered" not in st.session_state:
    st.session_state.overdue_email_triggered = False
if "nearby_email_trigger" not in st.session_state:
    st.session_state.nearby_email_trigger = [0, 0, 0, 0, 0, 0, 0]
if "last_warehouse" not in st.session_state:
    st.session_state.last_warehouse = None
if "map_date_filter" not in st.session_state:
    st.session_state.map_date_filter = []
if "action_plan_threshold_miles" not in st.session_state:
    st.session_state.action_plan_threshold_miles = 5
if "delivery_map_threshold_miles" not in st.session_state:
    st.session_state.delivery_map_threshold_miles = 5
if "b_name_choice" not in st.session_state:
    st.session_state.b_name_choice = []
if "c_name_choice" not in st.session_state:
    st.session_state.c_name_choice = []

# ─── UPLOAD DATA ───────────────────────────────────────────────────────────────
with st.expander("📂 Upload Data Files", expanded=True):
    col_upload1, col_upload2, col_upload3 = st.columns(3)
    with col_upload1:
        st.markdown('<div style="font-weight: bold; font-size: 16px;">Upload Dispatch file from M3/AHS110/AB-AB View (Max 15 Days) | <a href="https://mingle-portal.inforcloudsuite.com/PVSCHEMICALS_PRD/4e6b3a8f-6cb4-46a2-87ba-d0d88777a4b8?favoriteContext=bookmark?CMS100%26fieldNames=WFSLCT%252C071725%252CWTSLCT%252C071725%252CWFSLC2%252C%252520%252CWTSLC2%252C%252520%252CW1OBKV%252C%252520%252CW2OBKV%252C%252520%26tableName=CSYIBC%26keys=CKCONO%252C100%252CCKIBCA%252C%252B%26description=CMS100%2520Custom%2520list%26includeStartPanel=True%26source=MForms%26requirePanel=True%26startPanel=B%26sortingOrder=1%26view=STD02-AB%26parameters=XXIBCA%252C%2526%2526PVS_DISPATCH&LogicalId=lid://infor.m3.m3" target="_blank">Link</a></div>', unsafe_allow_html=True)
        uploaded_deliveries = st.file_uploader("Dispatch File", type=["xlsx", "xls"], label_visibility="collapsed", key="dispatch_uploader")
    with col_upload2:
        st.markdown('<div style="font-weight: bold; font-size: 16px;">Upload Stock Balance Excel file from AHS110/PVS_INVENTORY | <a href="https://mingle-portal.inforcloudsuite.com/PVSCHEMICALS_PRD/4e6b3a8f-6cb4-46a2-87ba-d0d88777a4b8?favoriteContext=bookmark?CMS100%26fieldNames=WWNFTR%252C0%252CWWAGGR%252C0%252CWWSLF1%252CMBWHLO%252CWWSLF2%252CMBITNO%252CWWSLF3%252CMBIPLA%252CWWSUB1%252C0%252CWWSUB2%252C0%252CWWSUB3%252C0%252CWFSLCT%252C%252520%252CWTSLCT%252C%252520%252CWFSLC2%252C100000%252CWTSLC2%252C99999%252CWFSLC3%252C30%252CWTSLC3%252C30%26tableName=CSYIBC%26keys=CKCONO%252C100%252CCKIBCA%252C%252B%26description=CMS100%2520Ad%2520hoc%2520report%2520designer%26includeStartPanel=True%26source=MForms%26requirePanel=True%26startPanel=B%26sortingOrder=1%26view=STD01-01%26parameters=XXIBCA%252C%2526%2526PVS_INVENTORY&LogicalId=lid://infor.m3.m3" target="_blank">Link</a></div>', unsafe_allow_html=True)
        uploaded_inventory = st.file_uploader("Inventory File", type=["xlsx", "xls"], label_visibility="collapsed", key="inventory_uploader")
    with col_upload3:
        st.markdown('<div style="font-weight: bold; font-size: 16px;">Upload the Overdue Containers file (Ex: 60/90+ days overdue) using this <a href="http://cpsdb011/Reports/Pages/Report.aspx?ItemPath=%2fM3+Reports%2fCustomer%2fCustomer_Asset_ETL" target="_blank">Link</a></div>', unsafe_allow_html=True)
        uploaded_overdue_containers = st.file_uploader("Overdue Containers File", type=["xlsx", "xls"], accept_multiple_files=False, help="Upload an Excel file containing overdue container data.", label_visibility="collapsed", key="overdue_uploader")

if not uploaded_deliveries:
    st.stop()
df_deliveries = pd.read_excel(uploaded_deliveries)

if uploaded_inventory:
    st.session_state.inventory_df = pd.read_excel(uploaded_inventory)
    required_inventory_cols = ["Item number", "Whs", "On-hand approve", "Allocated qty"]
    for col in required_inventory_cols:
        if col not in st.session_state.inventory_df.columns:
            st.error(f"Missing required column in inventory file: {col}")
            st.session_state.inventory_df = pd.DataFrame()
            st.stop()
    if not st.session_state.inventory_df.empty:
        st.session_state.inventory_df["Item number"] = st.session_state.inventory_df["Item number"].astype(str)
        st.session_state.inventory_df["Whs"] = st.session_state.inventory_df["Whs"].astype(str)
else:
    st.info("Upload the inventory Excel file to check stock levels for early delivery candidates.")
    st.session_state.inventory_df = pd.DataFrame()

if uploaded_overdue_containers:
    st.session_state.overdue_containers_df = pd.read_excel(uploaded_overdue_containers, header=1)
    required_overdue_cols = ["Bill to Name", "Tote Number", "Last Sale Date", "CO Number"]
    for col in required_overdue_cols:
        if col not in st.session_state.overdue_containers_df.columns:
            st.error(f"Missing required column in overdue containers file: {col}")
            st.session_state.overdue_containers_df = pd.DataFrame()
            st.stop()
    if len(st.session_state.overdue_containers_df.columns) < 4:
        st.error("Overdue containers file does not have a warehouse column in column D.")
        st.session_state.overdue_containers_df = pd.DataFrame()
        st.stop()
    st.session_state.overdue_containers_df["Warehouse"] = st.session_state.overdue_containers_df.iloc[:, 3].astype(str)
else:
    st.info("Upload the overdue containers Excel file to view recovery details.")
    st.session_state.overdue_containers_df = pd.DataFrame()

# ─── VERIFY REQUIRED COLUMNS (DELIVERIES) ──────────────────────────────────────────────────
required_cols_deliveries = [
    "Cfdldt", "Address 1", "City", "Sta", "Whs",
    "Dmt", "Dtm", "Otp", "His", "Salespers",
    "Description 1", "U/M", "Order qty bU/M", "Gross wt", "Name", "Item no", "Resp", "CO no"
]
for col in required_cols_deliveries:
    if col not in df_deliveries.columns:
        st.error(f"Missing required column in deliveries file: {col}")
        st.stop()

df = df_deliveries.copy()

# Add Route column if it doesn't exist
if "Route" not in df.columns:
    df["Route"] = ""  # Initialize with empty strings

# ─── FILTERS ───────────────────────────────────────────────────────────────
with st.expander("🔍 Filters", expanded=True):
    st.markdown('<h2 class="filters-heading">Filters</h2>', unsafe_allow_html=True)

    col_whs, col_gap, col_date = st.columns([0.10, 0.03, 0.87])
    with col_whs:
        st.markdown('<div class="filters-label">Select Warehouse</div>', unsafe_allow_html=True)
        whs = st.selectbox(
            "Select Warehouse",
            [""] + sorted(df["Whs"].unique()),
            label_visibility="hidden",
            key="whs_select"
        )
    if not whs:
        st.info("Select a warehouse to continue.")
        st.stop()
    
    if st.session_state.last_warehouse != whs:
        st.session_state.trajectories = [
            {"coords": [], "driver": "", "fill_rate": "", "lbs_to_add": "", "url": "", "added": False}
            for _ in range(7)
        ]
        st.session_state.last_warehouse = whs

    df = df[df["Whs"] == whs]

    with col_date:
        df["Date"] = pd.to_datetime(df["Cfdldt"]).dt.strftime("%m-%d-%Y")
        date_vals = sorted(df["Date"].unique())
        st.markdown('<div class="filters-label">Delivery Dates:</div>', unsafe_allow_html=True)
        date_choice = st.multiselect(
            "Delivery Dates:",
            date_vals, default=date_vals,
            label_visibility="hidden",
            key="date_multiselect"
        )
        current_date_hash = pd.util.hash_pandas_object(pd.Series(date_choice)).sum()
        if "date_filter_hash" not in st.session_state or st.session_state.date_filter_hash != current_date_hash:
            st.session_state.date_filter_hash = current_date_hash
            st.rerun()
    df = df[df["Date"].isin(date_choice)]

    st.markdown("<br>", unsafe_allow_html=True)

    dmt_vals = sorted(df["Dmt"].unique())
    dtm_vals = sorted(df["Dtm"].unique())
    otp_vals = sorted(df["Otp"].unique())
    his_vals = sorted([x for x in df["His"].unique() if str(x).isdigit() and int(x) <= 66])

    col_dmt, col_dtm, col_otp, col_status = st.columns(4)
    with col_dmt:
        st.markdown('<div class="filters-label">Delivery Method:</div>', unsafe_allow_html=True)
        default_dmt = []
        if "PVS" in dmt_vals:
            default_dmt.append("PVS")
        if "PVB" in dmt_vals:
            default_dmt.append("PVB")
        dmt_choice = st.multiselect(
            "Delivery Method:",
            dmt_vals, default=default_dmt,
            label_visibility="hidden",
            key="dmt_multiselect"
        )
    with col_dtm:
        st.markdown('<div class="filters-label">Delivery Terms:</div>', unsafe_allow_html=True)
        default_dtm = ["DEL"] if "DEL" in dtm_vals else []
        dtm_choice = st.multiselect(
            "Delivery Terms:",
            dtm_vals, default=default_dtm,
            label_visibility="hidden",
            key="dtm_multiselect"
        )
    with col_otp:
        st.markdown('<div class="filters-label">Order Type:</div>', unsafe_allow_html=True)
        default_otp = []
        if "REV" in otp_vals:
            default_otp.append("REV")
        if "DIR" in otp_vals:
            default_otp.append("DIR")
        otp_choice = st.multiselect(
            "Order Type:",
            otp_vals, default=default_otp,
            label_visibility="hidden",
            key="otp_multiselect"
        )
    with col_status:
        st.markdown('<div class="filters-label">Line Status:</div>', unsafe_allow_html=True)
        his_choice = st.multiselect(
            "Line Status:",
            his_vals, default=his_vals,
            label_visibility="hidden",
            key="his_multiselect"
        )
    df = df[df["Dmt"].isin(dmt_choice) & df["Dtm"].isin(dtm_choice)]
    df = df[df["Otp"].isin(otp_choice) & df["His"].isin(his_choice)]

    st.markdown("<br>", unsafe_allow_html=True)

current_df_hash = pd.util.hash_pandas_object(df, index=True).sum()

show_geocode_button = not st.session_state.run or (current_df_hash != st.session_state.last_filtered_df_hash)

if show_geocode_button:
    if st.button("Start Geocoding & Plotting"):
        st.session_state.df_geo, st.session_state.cmap, st.session_state.raw_ap_df, cache = \
            generate_initial_plan_data(df, st.session_state.inventory_df, cache)
        st.session_state.run = True
        st.session_state.last_filtered_df_hash = current_df_hash
        st.session_state.map_date_filter = sorted(st.session_state.df_geo["Date"].unique())
        st.session_state.b_name_choice = sorted(st.session_state.raw_ap_df["B Name"].dropna().unique())
        st.session_state.c_name_choice = sorted(st.session_state.raw_ap_df["C Name"].dropna().unique())
        st.rerun()

if st.session_state.run and not show_geocode_button:
    df_geo = st.session_state.df_geo
    cmap = st.session_state.cmap
    raw_ap_df = st.session_state.raw_ap_df

    # Get current time and date for file naming
    current_time = datetime.now().strftime("%I.%M%p").lstrip("0").replace("AM", "AM").replace("PM", "PM")
    current_date = datetime.now().strftime("%m-%d-%Y")


    # ...existing code...

    with st.expander("🗺️ Route-Based Delivery Map", expanded=False):
        st.markdown("### Route-Based Delivery Trajectories by Date")
        st.markdown("This map shows deliveries grouped by Route column, with each route representing a driver trajectory.")
        
        # Date selector for route map
        route_map_dates = sorted(df_geo["Date"].unique(), key=lambda x: pd.to_datetime(x, format="%m-%d-%Y", errors="coerce"))
        
        col_date, col_options = st.columns([1, 1])
        
        with col_date:
            selected_route_date = st.selectbox(
                "Select Date for Route Map",
                route_map_dates,
                key="route_map_date_select"
            )
        
        with col_options:
            st.markdown('<div style="margin-top: 8px;"></div>', unsafe_allow_html=True)
        
        if selected_route_date:
            # Filter data for selected date
            route_date_data = df_geo[df_geo["Date"] == selected_route_date].copy()
            
            if not route_date_data.empty:
                # Prepare detailed data with driver assignments first
                detailed_dispatch_data = route_date_data.copy()
                detailed_dispatch_data = detailed_dispatch_data.rename(columns={"Route": "Route Assignment"})
                if "Qty" in detailed_dispatch_data.columns and "Gross wt" in detailed_dispatch_data.columns:
                    detailed_dispatch_data["Total Weight (lbs)"] = (detailed_dispatch_data["Qty"] * detailed_dispatch_data["Gross wt"]).round(1)

                # Add Drivers column with default assignments
                unique_routes = detailed_dispatch_data["Route Assignment"].dropna().unique().tolist()
                unique_routes = [r for r in unique_routes if r != ""]
                route_number_map = {route: i+1 for i, route in enumerate(sorted(unique_routes))}
                detailed_dispatch_data["Drivers"] = detailed_dispatch_data["Route Assignment"].map(route_number_map).fillna(1)
                
                # Handle bulk deliveries (LB units with empty routes) - each gets unique driver
                bulk_mask = (
                    (detailed_dispatch_data["Route Assignment"].isna() | (detailed_dispatch_data["Route Assignment"] == "")) &
                    (detailed_dispatch_data["U/M"].str.upper() == "LB")
                )
                bulk_indices = detailed_dispatch_data[bulk_mask].index
                
                if len(bulk_indices) > 0:
                    # Start numbering bulk deliveries after the last routed driver
                    max_routed_driver = detailed_dispatch_data[~bulk_mask]["Drivers"].max() if not detailed_dispatch_data[~bulk_mask].empty else 0
                    next_driver = int(max_routed_driver) + 1 if pd.notna(max_routed_driver) else 1
                    
                    # Assign unique driver to each bulk delivery
                    for idx in bulk_indices:
                        detailed_dispatch_data.loc[idx, "Drivers"] = next_driver
                        next_driver += 1
                
                # Helper function to normalize customer names for priority assignment
                def normalize_customer_name(name):
                    """Normalize customer name to ensure consistent matching"""
                    import re
                    normalized = str(name).strip().upper()
                    # Remove extra spaces
                    normalized = re.sub(r'\s+', ' ', normalized)
                    # Remove common punctuation that might vary
                    normalized = re.sub(r'[.,\-\'"]', '', normalized)
                    return normalized
                
                # Place Drivers column before Route Assignment
                cols = detailed_dispatch_data.columns.tolist()
                if "Drivers" in cols and "Route Assignment" in cols:
                    cols.insert(cols.index("Route Assignment"), cols.pop(cols.index("Drivers")))
                detailed_dispatch_data = detailed_dispatch_data[cols]

                # Select columns to display
                priority_columns = [
                    "Date", "Drivers", "Priority", "Route Assignment", "Description 1", "Name", "full_address",
                    "His", "Line no", "CO no", "Item no", "U/M", "Order qty bU/M", "Gross wt", "Total Weight (lbs)"
                ]
                available_detailed_columns = []
                for col in priority_columns:
                    if col in detailed_dispatch_data.columns:
                        available_detailed_columns.append(col)
                # Exclude Address 1, City, and Sta since they're already in full_address
                # Also exclude Cfdldt since it's already shown as formatted "Date" column
                # Exclude redundant columns after Geo code Y
                excluded_columns = ["lat", "lon", "lat_jitter", "lon_jitter", "color", "Address 1", "City", "Sta", "Cfdldt", 
                                    "Product", "Qty", "Whs", "Resp", "Salespers", "Dmt", "Dtm", "Otp", "Delivery_Order"]
                for col in detailed_dispatch_data.columns:
                    if col not in available_detailed_columns and col not in excluded_columns:
                        available_detailed_columns.append(col)

                # Sort by driver number first, then by optimized delivery order
                # Add delivery order for each driver group
                detailed_dispatch_data_sorted = []
                
                for driver_num in sorted(detailed_dispatch_data["Drivers"].unique()):
                    driver_data = detailed_dispatch_data[detailed_dispatch_data["Drivers"] == driver_num].copy()
                    
                    if not driver_data.empty and whs in WAREHOUSE_COORDS:
                        # Get warehouse coordinates
                        warehouse_coords = [WAREHOUSE_COORDS[whs]["lat"], WAREHOUSE_COORDS[whs]["lon"]]
                        
                        # Optimize delivery order for this driver
                        optimized_driver_data = optimize_delivery_order(driver_data, warehouse_coords)
                        
                        # Add delivery sequence number
                        optimized_driver_data = optimized_driver_data.reset_index(drop=True)
                        optimized_driver_data["Delivery_Order"] = range(1, len(optimized_driver_data) + 1)
                        
                        detailed_dispatch_data_sorted.append(optimized_driver_data)
                    else:
                        # Fallback to original order if optimization fails
                        driver_data["Delivery_Order"] = range(1, len(driver_data) + 1)
                        detailed_dispatch_data_sorted.append(driver_data)
                
                if detailed_dispatch_data_sorted:
                    detailed_dispatch_data = pd.concat(detailed_dispatch_data_sorted, ignore_index=True)
                else:
                    detailed_dispatch_data["Delivery_Order"] = range(1, len(detailed_dispatch_data) + 1)
                
                # Add Priority column based on OPTIMIZED delivery order within each driver group
                detailed_dispatch_data["Priority"] = 0
                for driver_num in detailed_dispatch_data["Drivers"].unique():
                    driver_mask = detailed_dispatch_data["Drivers"] == driver_num
                    driver_data = detailed_dispatch_data[driver_mask].copy()
                    
                    # Assign priorities based on unique company locations in OPTIMIZED order
                    # First pass: identify all unique customers and assign them priority numbers
                    driver_data_sorted = driver_data.sort_values("Delivery_Order")
                    
                    # Get unique customers in the order they first appear
                    unique_customers = []
                    for idx, row in driver_data_sorted.iterrows():
                        company_name = normalize_customer_name(row["Name"])
                        if company_name not in unique_customers:
                            unique_customers.append(company_name)
                    
                    # Create priority mapping for all customers
                    company_priority = {customer: i + 1 for i, customer in enumerate(unique_customers)}
                    
                    # Second pass: assign priorities to all rows based on customer
                    for idx, row in driver_data_sorted.iterrows():
                        company_name = normalize_customer_name(row["Name"])
                        detailed_dispatch_data.loc[idx, "Priority"] = company_priority[company_name]
                
                # Update available columns to include Priority
                if "Priority" not in available_detailed_columns:
                    # Insert Priority after Delivery_Order if it exists, otherwise after Drivers
                    if "Delivery_Order" in available_detailed_columns:
                        insert_idx = available_detailed_columns.index("Delivery_Order") + 1
                    elif "Drivers" in available_detailed_columns:
                        insert_idx = available_detailed_columns.index("Drivers") + 1
                    else:
                        insert_idx = 0
                    available_detailed_columns.insert(insert_idx, "Priority")
                
                # Sort the detailed data by Drivers and Priority for consistent display
                detailed_dispatch_data = detailed_dispatch_data.sort_values(["Drivers", "Priority", "Name"]).reset_index(drop=True)
                
                # AUTOMATIC ROUTE OPTIMIZATION - Run once on initial load for all drivers
                if "optimized_dispatch_data" not in st.session_state and whs in WAREHOUSE_COORDS:
                    warehouse_coords = [WAREHOUSE_COORDS[whs]["lat"], WAREHOUSE_COORDS[whs]["lon"]]
                    optimized_data_list = []
                    
                    for driver_num in detailed_dispatch_data["Drivers"].unique():
                        driver_data = detailed_dispatch_data[detailed_dispatch_data["Drivers"] == driver_num].copy()
                        
                        if not driver_data.empty:
                            # Optimize this driver's route
                            optimized_driver_data = optimize_delivery_order(driver_data, warehouse_coords)
                            optimized_driver_data = optimized_driver_data.reset_index(drop=True)
                            
                            # Build unique company list in the order they first appear in optimized route
                            unique_companies_in_order = []
                            for _, row in optimized_driver_data.iterrows():
                                company_name = normalize_customer_name(row["Name"])
                                if company_name not in unique_companies_in_order:
                                    unique_companies_in_order.append(company_name)
                            
                            # Create fresh priority mapping: each unique company gets unique priority
                            company_priority_map = {company: idx + 1 for idx, company in enumerate(unique_companies_in_order)}
                            
                            # Assign priorities to optimized data
                            optimized_driver_data["Priority"] = optimized_driver_data["Name"].apply(
                                lambda name: company_priority_map[normalize_customer_name(name)]
                            )
                            
                            optimized_data_list.append(optimized_driver_data)
                    
                    if optimized_data_list:
                        optimized_full_data = pd.concat(optimized_data_list, ignore_index=True)
                        detailed_dispatch_data_reset = detailed_dispatch_data.copy()
                        
                        # Update ALL rows for each driver with correct optimized priorities
                        for driver_num in detailed_dispatch_data["Drivers"].unique():
                            driver_optimized_data = optimized_full_data[optimized_full_data["Drivers"] == driver_num].copy()
                            
                            if not driver_optimized_data.empty:
                                driver_optimized_data = driver_optimized_data.reset_index(drop=True)
                                
                                # Build unique company list in the order they first appear in optimized route
                                unique_companies_in_order = []
                                for _, row in driver_optimized_data.iterrows():
                                    company_name = normalize_customer_name(row["Name"])
                                    if company_name not in unique_companies_in_order:
                                        unique_companies_in_order.append(company_name)
                                
                                # Create fresh priority mapping: each unique company gets unique priority
                                company_priority_map = {company: idx + 1 for idx, company in enumerate(unique_companies_in_order)}
                                
                                # Update ALL rows in detailed_dispatch_data for this driver with correct priorities
                                for company_name, priority in company_priority_map.items():
                                    company_mask = (
                                        (detailed_dispatch_data_reset["Drivers"] == driver_num) &
                                        (detailed_dispatch_data_reset["Name"].apply(normalize_customer_name) == company_name)
                                    )
                                    detailed_dispatch_data_reset.loc[company_mask, "Priority"] = priority
                        
                        # Store optimized data in session state
                        st.session_state.optimized_dispatch_data = detailed_dispatch_data_reset
                        detailed_dispatch_data = detailed_dispatch_data_reset
                
                # Get the max driver number for dropdown options (include bulk deliveries)
                current_max_driver = detailed_dispatch_data["Drivers"].max() if not detailed_dispatch_data.empty else 1
                max_driver_num = max(int(current_max_driver), 10)
                
                # Display editable table for driver assignments
                st.markdown("### Assign Drivers and View Details")
                st.markdown("Edit the Driver numbers below and the map will update accordingly")
                st.info("🚛 **Route Optimization:** Driver routes are automatically optimized using real road networks for efficient delivery sequencing.")
                
                # Use optimized data if available, otherwise use regular detailed_dispatch_data
                data_for_editor = detailed_dispatch_data
                if "optimized_dispatch_data" in st.session_state:
                    data_for_editor = st.session_state.optimized_dispatch_data
                
                # Create unique key for data editor to force refresh when optimized
                editor_key = "driver_assignments"
                if "optimization_counter" in st.session_state:
                    editor_key = f"driver_assignments_{st.session_state.optimization_counter}"
                
                edited_data = st.data_editor(
                    data_for_editor[available_detailed_columns],
                    hide_index=True,
                    use_container_width=True,
                    height=800,
                    column_config={
                        "Drivers": st.column_config.SelectboxColumn(
                            "Drivers", 
                            help="Select driver number for this delivery",
                            options=list(range(1, max_driver_num + 1)),
                            required=True
                        ),
                        "Delivery_Order": st.column_config.NumberColumn("Delivery Order", help="Optimized delivery sequence for this driver"),
                        "Priority": st.column_config.NumberColumn(
                            "Priority", 
                            help="Stop priority - change to adjust delivery order",
                            min_value=1,
                            max_value=50,
                            step=1
                        ),
                        "Route Assignment": st.column_config.TextColumn("Route Assignment", help="Route assigned to delivery"),
                        "Description 1": st.column_config.TextColumn("Description 1", help="Product description"),
                        "Date": st.column_config.TextColumn("Dispatch Date", help="Delivery date"),
                        "Name": st.column_config.TextColumn("Customer Name", help="Customer company name"),
                        "full_address": st.column_config.TextColumn("Full Address", help="Complete delivery address"),
                        "His": st.column_config.TextColumn("His", help="Line status"),
                        "Line no": st.column_config.TextColumn("Line", help="Line number"),
                        "U/M": st.column_config.TextColumn("Unit", help="Unit of measure"),
                        "Order qty bU/M": st.column_config.NumberColumn("Order qty bU/M", help="Order quantity"),
                        "Gross wt": st.column_config.NumberColumn("Gross wt", help="Gross weight per unit"),
                        "Total Weight (lbs)": st.column_config.NumberColumn("Total Weight (lbs)", help="Calculated total weight"),
                        "CO no": st.column_config.TextColumn("CO Number", help="Customer order number"),
                        "Item no": st.column_config.TextColumn("Item Number", help="Product item number"),
                        "Resp": st.column_config.TextColumn("Responsible", help="Responsible person"),
                        "Salespers": st.column_config.TextColumn("Salesperson", help="Sales representative")
                    },
                    disabled=[col for col in available_detailed_columns if col not in ["Drivers", "Priority"]],
                    key=editor_key
                )

                # Detect changes in the edited data and trigger refresh
                current_data_hash = pd.util.hash_pandas_object(edited_data[["Drivers", "Priority"]]).sum()
                
                # Initialize or check for data changes
                if "previous_data_hash" not in st.session_state:
                    st.session_state.previous_data_hash = current_data_hash
                elif st.session_state.previous_data_hash != current_data_hash:
                    # Data has changed, update hash and trigger rerun for immediate refresh
                    st.session_state.previous_data_hash = current_data_hash
                    st.rerun()

                # Auto-sort the data by Drivers and Priority for better organization
                if "Priority" in edited_data.columns and "Drivers" in edited_data.columns:
                    edited_data = edited_data.sort_values(["Drivers", "Priority", "Name"]).reset_index(drop=True)

                # Add manual refresh button for immediate update
                col_refresh, col_info = st.columns([1, 4])
                with col_refresh:
                    if st.button("🔄 Refresh Table & Map", help="Click to update table sorting and map routes based on your priority changes"):
                        st.rerun()
                with col_info:
                    st.info("💡 Table automatically resorts and map updates when you change priorities. Use refresh button if needed.")

                # Display driver statistics summary right after the table
                st.markdown("### Driver Summary")
                col_minutes, col_coefficient = st.columns(2)
                with col_minutes:
                    minutes_per_stop = st.number_input(
                        "Minutes per stop:",
                        min_value=1,
                        max_value=120,
                        value=30,
                        step=5,
                        help="Estimated time (in minutes) spent at each customer stop"
                    )
                with col_coefficient:
                    travel_time_coefficient = st.number_input(
                        "Travel time coefficient:",
                        min_value=1.0,
                        max_value=2.0,
                        value=1.2,
                        step=0.1,
                        format="%.1f",
                        help="Multiplier for travel time to account for truck/trailer vs. car (1.2 = 20% slower)"
                    )
                summary_data = []
                total_deliveries = 0
                total_customers = 0
                total_weight_all = 0
                total_miles_all = 0
                total_time_all = 0
                
                # We'll populate this after driver_stats is calculated below
                # Placeholder - will be filled after map generation
                summary_placeholder = st.empty()

                # Now create map based on driver assignments from edited data
                # Merge edited driver assignments with original coordinate data
                # Create a mapping of driver assignments and priorities based on row index
                driver_assignments = edited_data["Drivers"].fillna(0).astype(int)
                
                # Handle Priority column - it may not exist in initial load
                if "Priority" in edited_data.columns:
                    priority_assignments = edited_data["Priority"].fillna(1).astype(int)
                else:
                    # Use the already calculated priorities from detailed_dispatch_data
                    if "Priority" in detailed_dispatch_data.columns:
                        priority_assignments = detailed_dispatch_data["Priority"].fillna(1).astype(int)
                    else:
                        # Ultimate fallback: create sequential priorities
                        priority_assignments = pd.Series(range(1, len(edited_data) + 1), index=edited_data.index)
                
                route_date_data_with_drivers = route_date_data.copy()
                # Reset index to ensure alignment
                route_date_data_with_drivers = route_date_data_with_drivers.reset_index(drop=True)
                driver_assignments_reset = driver_assignments.reset_index(drop=True)
                priority_assignments_reset = priority_assignments.reset_index(drop=True)
                
                # Only assign if lengths match
                if len(route_date_data_with_drivers) == len(driver_assignments_reset):
                    route_date_data_with_drivers["Driver_Group"] = driver_assignments_reset
                    route_date_data_with_drivers["Priority"] = priority_assignments_reset
                else:
                    # Fallback: use original detailed_dispatch_data
                    route_date_data_with_drivers["Driver_Group"] = detailed_dispatch_data["Drivers"].reset_index(drop=True)
                    route_date_data_with_drivers["Priority"] = detailed_dispatch_data["Priority"].reset_index(drop=True) if "Priority" in detailed_dispatch_data.columns else 1
                
                unique_drivers = sorted([d for d in route_date_data_with_drivers["Driver_Group"].unique() if d > 0])
                
                # Driver selection checkboxes
                st.markdown("### Select Drivers to Display")
                if unique_drivers:
                    # Initialize driver selection state if not exists
                    if "selected_drivers_state" not in st.session_state:
                        st.session_state.selected_drivers_state = {driver: True for driver in unique_drivers}
                    
                    # Ensure all current drivers exist in selection state
                    for driver in unique_drivers:
                        if driver not in st.session_state.selected_drivers_state:
                            st.session_state.selected_drivers_state[driver] = True
                    
                    # Limit to max 10 columns per row to prevent layout issues
                    drivers_per_row = min(10, len(unique_drivers))
                    selected_drivers = []
                    
                    # Create rows of checkboxes
                    for row_start in range(0, len(unique_drivers), drivers_per_row):
                        row_end = min(row_start + drivers_per_row, len(unique_drivers))
                        drivers_in_row = unique_drivers[row_start:row_end]
                        col_drivers = st.columns(len(drivers_in_row))
                        
                        for i, driver_num in enumerate(drivers_in_row):
                            with col_drivers[i]:
                                # Use session state to preserve selection across refreshes
                                current_value = st.session_state.selected_drivers_state.get(driver_num, True)
                                if st.checkbox(f"Driver {driver_num}", value=current_value, key=f"driver_select_{driver_num}"):
                                    selected_drivers.append(driver_num)
                                    st.session_state.selected_drivers_state[driver_num] = True
                                else:
                                    st.session_state.selected_drivers_state[driver_num] = False
                else:
                    selected_drivers = []
                    st.info("No drivers found for this date")
                
                # Debug: Show selected drivers count
                if selected_drivers:
                    st.caption(f"✓ {len(selected_drivers)} driver(s) selected: {', '.join([f'Driver {d}' for d in sorted(selected_drivers)])}")
                
                # Create color mapping for drivers (supports up to 15 unique drivers)
                driver_colors = [
                    'blue', 'red', 'green', 'purple', 'orange', 
                    'darkred', 'darkgreen', 'cadetblue', 'darkblue', 'lightcoral',
                    'pink', 'lightblue', 'lightgreen', 'beige', 'gray'
                ]
                driver_color_map = {driver: driver_colors[i % len(driver_colors)] for i, driver in enumerate(unique_drivers)}
                
                # Create driver map
                if not edited_data.empty:
                    # Use the original route_date_data for center coordinates since edited_data doesn't have lat_jitter
                    center_lat = route_date_data["lat_jitter"].mean()
                    center_lon = route_date_data["lon_jitter"].mean()
                else:
                    center_lat = WAREHOUSE_COORDS.get(whs, {"lat": 40.7128}).get("lat", 40.7128)
                    center_lon = WAREHOUSE_COORDS.get(whs, {"lon": -74.0060}).get("lon", -74.0060)

                driver_map = folium.Map(
                    location=[center_lat, center_lon],
                    zoom_start=10,
                    tiles="OpenStreetMap",
                    zoom_control=True
                )

                # Add warehouse marker
                if whs in WAREHOUSE_COORDS:
                    warehouse = WAREHOUSE_COORDS[whs]
                    folium.Marker(
                        location=[warehouse["lat"], warehouse["lon"]],
                        tooltip=folium.Tooltip(f"<b>{warehouse['name']}</b>"),
                        icon=folium.Icon(color="red", icon="star", prefix="fa")
                    ).add_to(driver_map)

                # Add delivery points and create routes for each driver group
                driver_stats = {}
                for driver_num in unique_drivers:
                    # Process ALL drivers for stats, but only add map markers for selected drivers
                    add_to_map = driver_num in selected_drivers
                        
                    # Use the exact edited data to ensure consistency
                    edited_driver_data = edited_data[edited_data["Drivers"] == driver_num].copy()
                    
                    # Merge with original data to get coordinates, but only for rows that exist in edited_data
                    driver_data_list = []
                    for idx, edited_row in edited_driver_data.iterrows():
                        # Find matching row in original data based on key fields that exist in edited_data
                        matching_rows = route_date_data_with_drivers[
                            (route_date_data_with_drivers["Name"] == edited_row["Name"]) & 
                            (route_date_data_with_drivers["CO no"] == edited_row["CO no"]) &
                            (route_date_data_with_drivers["Item no"] == edited_row["Item no"]) &
                            (route_date_data_with_drivers["Date"] == edited_row["Date"])
                        ]
                        
                        if not matching_rows.empty:
                            # Take the first match and combine with edited data
                            original_row = matching_rows.iloc[0].copy()
                            # Update with edited values
                            for col in edited_row.index:
                                if col in original_row.index:
                                    original_row[col] = edited_row[col]
                            driver_data_list.append(original_row)
                    
                    if driver_data_list:
                        driver_data = pd.DataFrame(driver_data_list).reset_index(drop=True)
                    else:
                        # Fallback: use edited data without coordinates
                        driver_data = edited_driver_data.copy()
                    
                    driver_color = driver_color_map[driver_num]
                    
                    # Sort by user-defined priority instead of auto-optimization
                    driver_data = driver_data.sort_values(["Priority", "Name"]).reset_index(drop=True)
                    
                    # Calculate driver statistics
                    total_weight = 0
                    for _, row in driver_data.iterrows():
                        if pd.notna(row["Qty"]) and pd.notna(row["Gross wt"]):
                            total_weight += row["Qty"] * row["Gross wt"]
                    
                    delivery_count = len(driver_data)
                    customer_count = driver_data["Name"].nunique()
                    
                    # Add delivery markers for this driver and build coordinates list
                    driver_coords = []
                    markers_added = 0
                    
                    # Use user-defined priorities from the edited data
                    # Track which priorities we've already added markers for
                    priority_markers_added = set()
                    
                    for idx, row in driver_data.iterrows():
                        if pd.notna(row["lat_jitter"]) and pd.notna(row["lon_jitter"]):
                            # Get priority from user-edited data
                            priority = row["Priority"]
                            
                            # Add delivery marker
                            try:
                                tooltip_text = (
                                    f"<b>Driver:</b> {driver_num}<br>"
                                    f"<b>Priority:</b> {priority}<br>"
                                    f"<b>Company:</b> {sanitize_for_html(row['Name'])}<br>"
                                    f"<b>Address:</b> {sanitize_for_html(row['full_address'])}<br>"
                                    f"<b>Product:</b> {sanitize_for_html(row['Product'])}<br>"
                                    f"<b>Qty:</b> {row['Qty']} {sanitize_for_html(row['U/M'])}<br>"
                                    f"<b>Weight:</b> {round(row['Qty'] * row['Gross wt'], 1) if pd.notna(row['Qty']) and pd.notna(row['Gross wt']) else 'N/A'} lbs<br>"
                                    f"<b>CO no:</b> {sanitize_for_html(row['CO no'])}<br>"
                                )
                                
                                tooltip = folium.Tooltip(tooltip_text, sticky=True)
                                
                                # Only add markers to map for selected drivers
                                if add_to_map:
                                    # All drivers get the same marker style (no "Unassigned" concept for drivers)
                                    folium.CircleMarker(
                                        location=(row["lat_jitter"], row["lon_jitter"]),
                                        radius=8,
                                        stroke=True,
                                        color='white',
                                        weight=2,
                                        fill=True,
                                        fill_opacity=0.8,
                                        fill_color=driver_color,
                                        tooltip=tooltip
                                    ).add_to(driver_map)
                                
                                # Add priority number marker (only once per priority)
                                if add_to_map and priority not in priority_markers_added:
                                    # Collect all deliveries for this priority to show in tooltip
                                    priority_deliveries = driver_data[driver_data["Priority"] == priority]
                                    
                                    # Build tooltip with priority details and all deliveries
                                    priority_tooltip_text = (
                                        f"<b>Priority:</b> {priority}<br>"
                                        f"<b>Company:</b> {sanitize_for_html(row['Name'])}<br>"
                                        f"<b>Address:</b> {sanitize_for_html(row['full_address'])}<br>"
                                        f"<b>Deliveries:</b><br>"
                                    )
                                    
                                    for _, delivery in priority_deliveries.iterrows():
                                        weight = round(delivery['Qty'] * delivery['Gross wt'], 1) if pd.notna(delivery['Qty']) and pd.notna(delivery['Gross wt']) else 'N/A'
                                        priority_tooltip_text += (
                                            f"• {sanitize_for_html(delivery['Product'])} - "
                                            f"{delivery['Qty']} {sanitize_for_html(delivery['U/M'])} ({weight} lbs)<br>"
                                        )
                                    
                                    priority_tooltip = folium.Tooltip(priority_tooltip_text, sticky=True)
                                    
                                    # Offset the priority marker slightly to avoid overlap with delivery marker
                                    offset_lat = row["lat_jitter"] + 0.001  # Small offset
                                    offset_lon = row["lon_jitter"] + 0.001
                                    
                                    folium.Marker(
                                        location=(offset_lat, offset_lon),
                                        tooltip=priority_tooltip,
                                        icon=folium.DivIcon(
                                            icon_size=(24, 24),
                                            icon_anchor=(12, 12),
                                            html=f'<div style="background-color:white;color:{driver_color};border:2px solid {driver_color};border-radius:50%;text-align:center;font-weight:bold;font-size:14px;line-height:20px;width:24px;height:24px;">{priority}</div>'
                                        )
                                    ).add_to(driver_map)
                                    priority_markers_added.add(priority)
                                
                                driver_coords.append([row["lat_jitter"], row["lon_jitter"]])
                                if add_to_map:
                                    markers_added += 1
                            except Exception:
                                continue
                    
                    # Calculate total miles and travel time using Google Maps routing (after driver_coords is built)
                    total_miles = 0
                    total_travel_time = 0  # in minutes
                    if len(driver_coords) >= 1 and whs in WAREHOUSE_COORDS:
                        try:
                            warehouse_coords = [WAREHOUSE_COORDS[whs]["lat"], WAREHOUSE_COORDS[whs]["lon"]]
                            
                            # Prepare addresses for distance calculation
                            addresses = []
                            addresses.append(f"{warehouse_coords[0]},{warehouse_coords[1]}")
                            for coord in driver_coords:
                                addresses.append(f"{coord[0]},{coord[1]}")
                            addresses.append(f"{warehouse_coords[0]},{warehouse_coords[1]}")
                            
                            # Get route distance and time from Google Maps with truck-specific parameters
                            if len(addresses) >= 2:
                                waypoints = addresses[1:-1] if len(addresses) > 2 else []
                                directions_result = gmaps_client.directions(
                                    origin=addresses[0],
                                    destination=addresses[-1],
                                    waypoints=waypoints,
                                    mode="driving",
                                    optimize_waypoints=False,  # Keep our priority-based order
                                    avoid=["tolls"],  # Trucks often avoid tolls
                                    departure_time="now"  # Get current traffic conditions
                                )
                                
                                if directions_result and 'legs' in directions_result[0]:
                                    # Sum up all leg distances and durations
                                    total_meters = sum(leg['distance']['value'] for leg in directions_result[0]['legs'])
                                    total_seconds = sum(leg['duration']['value'] for leg in directions_result[0]['legs'])
                                    
                                    total_miles = round(total_meters * 0.000621371, 1)  # Convert meters to miles
                                    total_travel_time = round(total_seconds / 60, 0)  # Convert seconds to minutes
                        except Exception as e:
                            # Fallback: estimate using straight-line distance * 1.3 (typical road factor)
                            if len(driver_coords) >= 1:
                                from geopy.distance import geodesic
                                estimated_distance = 0
                                prev_point = warehouse_coords
                                for coord in driver_coords:
                                    estimated_distance += geodesic(prev_point, coord).miles
                                    prev_point = coord
                                # Add return to warehouse
                                estimated_distance += geodesic(prev_point, warehouse_coords).miles
                                total_miles = round(estimated_distance * 1.3, 1)  # Apply road factor
                                
                                # Estimate travel time: assume 35 mph average for trucks in mixed traffic
                                total_travel_time = round((total_miles / 35) * 60, 0)  # Convert to minutes
                    
                    driver_stats[f"Driver {driver_num}"] = {
                        "deliveries": delivery_count,
                        "customers": customer_count,
                        "total_weight": round(total_weight, 2),
                        "total_miles": total_miles,
                        "travel_time_minutes": int(total_travel_time),
                        "color": driver_color,
                        "markers_on_map": markers_added
                    }
                    
                    # Create trajectory for this driver
                    if len(driver_coords) >= 1 and add_to_map:
                        # Always connect to warehouse for complete trajectory
                        if whs in WAREHOUSE_COORDS:
                            warehouse_coords = [WAREHOUSE_COORDS[whs]["lat"], WAREHOUSE_COORDS[whs]["lon"]]
                            
                            # Prepare addresses for Google Maps routing
                            addresses = []
                            
                            # Start with warehouse
                            warehouse_address = f"{warehouse_coords[0]},{warehouse_coords[1]}"
                            addresses.append(warehouse_address)
                            
                            # Add all delivery addresses
                            for coord in driver_coords:
                                addresses.append(f"{coord[0]},{coord[1]}")
                            
                            # Return to warehouse
                            addresses.append(warehouse_address)
                            
                            # Get realistic driving route using Google Maps
                            try:
                                driving_coords = get_driving_route(addresses, gmaps_client)
                                
                                if driving_coords:
                                    # Use the road-based coordinates
                                    route_coordinates = driving_coords
                                else:
                                    # Fallback to straight lines if API fails
                                    route_coordinates = [warehouse_coords] + driver_coords + [warehouse_coords]
                                    
                            except Exception as e:
                                # Fallback to straight lines if there's an error
                                route_coordinates = [warehouse_coords] + driver_coords + [warehouse_coords]
                            
                            # Create polyline for the realistic driver route
                            travel_time_formatted = format_travel_time(total_travel_time)
                            folium.PolyLine(
                                locations=route_coordinates,
                                color=driver_color,
                                weight=4,
                                opacity=0.8,
                                tooltip=f"Driver {driver_num} | {delivery_count} deliveries | {customer_count} customers | {round(total_weight, 0):,} lbs | {total_miles:.1f} mi | {travel_time_formatted}",
                                popup=folium.Popup(
                                    f"<b>Driver:</b> {driver_num}<br>"
                                    f"<b>Deliveries:</b> {delivery_count}<br>"
                                    f"<b>Customers:</b> {customer_count}<br>"
                                    f"<b>Total Weight:</b> {round(total_weight, 0):,} lbs<br>"
                                    f"<b>Total Miles:</b> {total_miles:.1f} mi<br>"
                                    f"<b>Travel Time:</b> {travel_time_formatted}",
                                    max_width=220
                                )
                            ).add_to(driver_map)

                # Create legend for drivers
                legend_html = f"""
                  <div style="
                    position: fixed; top: 10px; left: 50px; width: 280px;
                    background: white; padding: 15px; border: 2px solid black;
                    z-index: 1000; font-size: 14px; max-height: 90vh; overflow-y: auto;
                  ">
                    <b><strong>Driver Legend - {selected_route_date}:</strong></b><br><br>
                """
                
                if whs in WAREHOUSE_COORDS:
                    legend_html += (
                        '<i style="color:red;font-family:FontAwesome;font-size:12px;margin-right:5px;">★</i> '
                        f'{sanitize_for_html(WAREHOUSE_COORDS[whs]["name"])}<br><br>'
                    )
                
                legend_html += "<b>Drivers:</b><br>"
                for driver_name, stats in driver_stats.items():
                    # Only show selected drivers in the legend
                    driver_num = int(driver_name.replace("Driver ", ""))
                    if driver_num not in selected_drivers:
                        continue
                    # Calculate total time same as driver summary table (Travel Time + Stop Time)
                    adjusted_travel_time = stats["travel_time_minutes"] * travel_time_coefficient
                    stop_time = stats["customers"] * minutes_per_stop
                    total_time = adjusted_travel_time + stop_time
                    legend_html += (
                        f'<div style="margin:2px 0;">'
                        f'<span style="background-color:{stats["color"]};color:white;padding:2px 6px;border-radius:3px;font-size:11px;font-weight:bold;margin-right:5px;">{sanitize_for_html(driver_name)}</span>'
                        f'<span style="font-size:11px;">{stats["deliveries"]} del | {stats["customers"]} cust | {stats["total_weight"]:,.0f} lbs | {stats["total_miles"]:.1f} mi | {format_travel_time(total_time)}</span>'
                        f'</div>'
                    )
                
                legend_html += "</div>"
                driver_map.get_root().html.add_child(folium.Element(legend_html))

                # Populate Driver Summary table now that driver_stats is available
                total_stop_time_all = 0
                for driver_name, stats in driver_stats.items():
                    # Apply travel time coefficient to adjust for truck/trailer vs. car
                    adjusted_travel_time = stats["travel_time_minutes"] * travel_time_coefficient
                    travel_time_formatted = format_travel_time(adjusted_travel_time)
                    stop_time_minutes = stats["customers"] * minutes_per_stop
                    stop_time_formatted = format_travel_time(stop_time_minutes)
                    total_time_minutes = adjusted_travel_time + stop_time_minutes
                    total_time_formatted = format_travel_time(total_time_minutes)
                    summary_data.append({
                        "Driver": driver_name,
                        "Customers/Stops": stats["customers"],
                        "# of Deliveries": stats["deliveries"],
                        "Total Weight (lbs)": f"{stats['total_weight']:,.0f}",
                        "Total Miles": f"{stats['total_miles']:.1f}",
                        "Travel Time": travel_time_formatted,
                        "Stop Time": stop_time_formatted,
                        "Total Time": total_time_formatted
                    })
                    total_deliveries += stats["deliveries"]
                    total_customers += stats["customers"]
                    total_weight_all += stats["total_weight"]
                    total_miles_all += stats["total_miles"]
                    total_time_all += adjusted_travel_time
                    total_stop_time_all += stop_time_minutes
                
                # Add total row
                total_time_formatted = format_travel_time(total_time_all)
                total_stop_time_formatted = format_travel_time(total_stop_time_all)
                grand_total_time_formatted = format_travel_time(total_time_all + total_stop_time_all)
                summary_data.append({
                    "Driver": "TOTAL",
                    "Customers/Stops": total_customers,
                    "# of Deliveries": total_deliveries,
                    "Total Weight (lbs)": f"{total_weight_all:,.0f}",
                    "Total Miles": f"{total_miles_all:.1f}",
                    "Travel Time": total_time_formatted,
                    "Stop Time": total_stop_time_formatted,
                    "Total Time": grand_total_time_formatted
                })
                
                # Display the summary table in the placeholder
                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                    with summary_placeholder.container():
                        st.dataframe(summary_df, hide_index=True, width='stretch')

                # Add export functionality for edited data
                if not edited_data.empty:
                    col1, col2 = st.columns([1, 3])
                    with col1:
                        if st.button("📊 Export Driver Data", key="export_driver_dispatch"):
                            excel_content = generate_excel_file(
                                edited_data, 
                                sheet_name=f"Drivers_{selected_route_date}", 
                                align_left=True
                            )
                            if excel_content:
                                file_name = f"{whs}_Driver_Assignments_{selected_route_date.replace('-', '')}_{current_time}.xlsx"
                                st.download_button(
                                    label="Download Driver Assignments Excel",
                                    data=excel_content,
                                    file_name=file_name,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="download_driver_assignments"
                                )

                # Display the map
                components.html(driver_map._repr_html_(), height=800)
                
                # Check for SELECTED drivers with missing map markers (exclude unselected drivers)
                drivers_without_markers = [
                    name for name, stats in driver_stats.items() 
                    if stats.get("markers_on_map", 0) == 0 
                    and int(name.replace("Driver ", "")) in selected_drivers
                ]
                if drivers_without_markers:
                    st.warning(f"⚠️ The following driver(s) have deliveries but no map markers (missing geocoded coordinates): {', '.join(drivers_without_markers)}")
                
            else:
                st.info(f"No delivery data available for {selected_route_date}")
        else:
            st.info("Please select a date to view driver-based delivery map")

    with st.expander("🚚 Add Planned Driving Trajectories", expanded=False):

        st.markdown('<div style="font-weight: bold; font-size: 18px;">Upload Driver and Trajectory File</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)  # Add space above radio buttons
        input_mode = st.radio(
            "Input Mode",
            ["Upload Excel File", "Paste Table Data"],
            index=0,
            horizontal=True,
            label_visibility="collapsed",
            key="trajectory_input_mode"
        )

        if input_mode == "Upload Excel File":
            uploaded_trajectory_file = st.file_uploader(
                "Upload Driver and Trajectory File (Excel with Driver Names in A1:A7, URLs in J1:J7, Fill Rates in D1:D7, Max LBS in B1:B7, Current LBS in C1:C7)",
                type=["xlsx", "xls"],
                label_visibility="collapsed",
                key="trajectory_file_uploader"
            )

            dispatch_date_excel = None
            dispatch_date_str = None
            if uploaded_trajectory_file:
                try:
                    wb = load_workbook(uploaded_trajectory_file, data_only=True)
                    ws = wb.active
                    # Extract dispatch date from B1 (between 'Dispatching for: ' and '|')
                    dispatching_for = ""
                    b1_val = ws['B1'].value
                    dispatch_date_str = ""
                    if b1_val:
                        b1_val_str = str(b1_val).strip()
                        # Find substring between 'Dispatching for: ' and '|'
                        start_tag = "Dispatching for: "
                        end_tag = "|"
                        start_idx = b1_val_str.find(start_tag)
                        end_idx = b1_val_str.find(end_tag, start_idx + len(start_tag))
                        if start_idx != -1 and end_idx != -1:
                            dispatch_date_raw = b1_val_str[start_idx + len(start_tag):end_idx].strip()
                            dispatching_for = dispatch_date_raw
                        else:
                            # Fallback: try to use everything before '|' or the whole cell
                            if '|' in b1_val_str:
                                dispatching_for = b1_val_str.split('|')[0].strip()
                            else:
                                dispatching_for = b1_val_str.strip()
                        # Normalize dispatch date: remove extra spaces, convert to datetime, format as MM-DD-YY
                        try:
                            dispatch_date_excel = pd.to_datetime(dispatching_for, errors="coerce")
                            if not pd.isna(dispatch_date_excel):
                                dispatch_date_str = dispatch_date_excel.strftime("%m-%d-%Y")
                            else:
                                dispatch_date_str = str(dispatching_for).strip()
                        except Exception:
                            dispatch_date_str = str(dispatching_for).strip()
                        # Remove extra spaces and ensure format is MM-DD-YY
                        dispatch_date_str = dispatch_date_str.replace(" ", "")
                    st.session_state.dispatching_for = dispatch_date_str
                    # Show Dispatch Date field after upload and before Planned Driving Trajectories
                    if dispatch_date_str:
                        # Custom CSS to decrease padding and set field width to 1/10 container
                        st.markdown("""
                            <style>
                            .dispatch-date-label-upload {
                                font-weight: bold;
                                font-size: 16px;
                                text-align:left;
                                margin-bottom:0px;
                                margin-top:10px;
                                padding-top:0px;
                            }
                            /* Only target the Dispatch Date field by its key */
                            div[data-testid='stTextInput'][data-testid-custom='dispatch_date_excel_display'] input[data-testid='stTextInput-input'] {
                                width: 10vw !important;
                                min-width: 80px !important;
                                max-width: 120px !important;
                            }
                            div[data-testid='stTextInput'][data-testid-custom='dispatch_date_excel_display'] {
                                width: 10vw !important;
                                min-width: 80px !important;
                                max-width: 120px !important;
                                display: inline-block !important;
                            }
                            </style>
                        """, unsafe_allow_html=True)
                        st.markdown('<div class="dispatch-date-label-upload">Dispatch Date</div>', unsafe_allow_html=True)
                        st.text_input("Dispatch Date", value=dispatch_date_str, disabled=True, key="dispatch_date_excel_display", label_visibility="hidden")
                    # Read driver data from row 3 (A3:J3, etc.)
                    traj_data = []
                    for row in range(3, 10):
                        driver = ws[f'A{row}'].value if ws[f'A{row}'].value else ""
                        max_lbs = float(ws[f'B{row}'].value) if ws[f'B{row}'].value and pd.notna(ws[f'B{row}'].value) else 0
                        current_lbs = float(ws[f'C{row}'].value) if ws[f'C{row}'].value and pd.notna(ws[f'C{row}'].value) else 0
                        fill_rate_value = ws[f'D{row}'].value
                        if fill_rate_value and pd.notna(fill_rate_value):
                            try:
                                fill_rate = f"{float(fill_rate_value) * 100:.0f}%"
                            except ValueError:
                                fill_rate = ""
                        else:
                            fill_rate = ""
                        url = ws[f'J{row}'].hyperlink.target if ws[f'J{row}'].hyperlink else ""
                        traj_data.append([driver, max_lbs, current_lbs, fill_rate, url])
                    traj_df = pd.DataFrame(traj_data, columns=['Driver', 'Max_LBS', 'Current_LBS', 'Fill_Rate', 'URL'])
                    
                    # Ensure trajectories list has exactly 7 elements
                    while len(st.session_state.trajectories) < 7:
                        st.session_state.trajectories.append({"coords": [], "driver": "", "fill_rate": "", "lbs_to_add": "", "url": "", "added": False})
                    
                    for i in range(7):
                        if i < len(traj_df):
                            driver = str(traj_df.iloc[i, 0]) if pd.notna(traj_df.iloc[i, 0]) else ""
                            url = str(traj_df.iloc[i, 4]) if pd.notna(traj_df.iloc[i, 4]) else ""
                            fill_rate = str(traj_df.iloc[i, 3]) if pd.notna(traj_df.iloc[i, 3]) else ""
                            max_lbs = float(traj_df.iloc[i, 1]) if pd.notna(traj_df.iloc[i, 1]) else 0
                            current_lbs = float(traj_df.iloc[i, 2]) if pd.notna(traj_df.iloc[i, 2]) else 0
                            lbs_to_add = max_lbs - current_lbs if max_lbs != 0 or current_lbs != 0 else ""
                            # Format LBS to add with thousand separator, no decimals
                            if isinstance(lbs_to_add, (int, float)):
                                lbs_to_add_fmt = f"{int(round(lbs_to_add)):,}"
                            else:
                                lbs_to_add_fmt = ""
                            st.session_state.trajectories[i].update({
                                "driver": driver,
                                "fill_rate": fill_rate,
                                "lbs_to_add": lbs_to_add_fmt,
                                "url": url
                            })
                    st.success("Driver and trajectory data loaded successfully.")
                except Exception as e:
                    st.error(f"Error reading trajectory file: {e}")
            else:
                # If file is removed, clear all trajectory fields, map, and per-driver tables
                st.session_state.trajectories = [
                    {"coords": [], "driver": "", "fill_rate": "", "lbs_to_add": "", "url": "", "added": False}
                    for _ in range(7)
                ]
                if hasattr(st.session_state, 'dispatching_for'):
                    del st.session_state.dispatching_for

        elif input_mode == "Paste Table Data":

            # Dispatch Date input as date picker, required (move above paste table)
            dispatch_date_key = "dispatching_for_paste"
            dispatch_date_default = st.session_state.get(dispatch_date_key, None)
            # Custom CSS to further decrease padding above Dispatch Date field and set field width
            st.markdown("""
                <style>
                .dispatch-date-label {
                    font-weight: bold;
                    font-size: 16px;
                    text-align:left;
                    margin-bottom:0px;
                    padding-bottom:-18px;
                    margin-top:0px;
                    padding-top:0px;
                }
                div[data-testid='stDateInput'] {
                    margin-top: -18px !important;
                    padding-top: 0px !important;
                    width: 10% !important;
                    min-width: 120px !important;
                }
            </style>
            """, unsafe_allow_html=True)
            st.markdown('<div class="dispatch-date-label">Dispatch Date</div>', unsafe_allow_html=True)
            dispatch_date = st.date_input(
                "Dispatch Date",
                value=dispatch_date_default if dispatch_date_default else None,
                key=dispatch_date_key,
                format="MM-DD-YYYY",
                label_visibility="hidden"
            )

            use_paste_disabled = dispatch_date is None
            if use_paste_disabled:
                st.warning("Please select a Dispatch Date before using pasted trajectory data.")

            st.markdown(
                '<div style="font-weight: bold; font-size: 16px;">'
                'Copy and paste data from '
                '<a href="https://pvschem.sharepoint.com/sites/PVSNolwoodDispatch/" target="_blank" style="color:#1976d2;text-decoration:underline;font-weight:bold;">Sharepoint Dispatch Site</a> '
                '(Driver/Trailer to Map Trajectory column)'
                '</div>',
                unsafe_allow_html=True
            )
            paste_columns = [
                "Driver/Trailer",
                "Allowed Weight",
                "Total Gross Weight",
                "Current Fill Rate",
                "Customers served",
                "Roundtrip Route Miles",
                "Total Driving Time",
                "Total Unload time",
                "Total Time",
                "Map Trajectory"
            ]
            # Always enforce the correct columns in the paste table
            if "trajectory_paste_df" not in st.session_state or list(st.session_state.trajectory_paste_df.columns) != paste_columns:
                st.session_state.trajectory_paste_df = pd.DataFrame([[""]*len(paste_columns) for _ in range(8)], columns=paste_columns)

            # Initialize editor version for force refreshes
            if "paste_editor_version" not in st.session_state:
                st.session_state.paste_editor_version = 0

            # Show the editor directly with session state data
            paste_df = st.data_editor(
                st.session_state.trajectory_paste_df,
                column_order=paste_columns,
                key=f"trajectory_paste_editor_v{st.session_state.paste_editor_version}",
                width='stretch',
                hide_index=True,
                num_rows="dynamic",
                disabled=False,
                column_config={
                    "Driver/Trailer": st.column_config.TextColumn(
                        "Driver/Trailer",
                        help="Paste or type driver/trailer information",
                        max_chars=50
                    ),
                    "Allowed Weight": st.column_config.TextColumn(
                        "Allowed Weight",
                        help="Weight capacity"
                    ),
                    "Total Gross Weight": st.column_config.TextColumn(
                        "Total Gross Weight",
                        help="Current weight"
                    ),
                    "Current Fill Rate": st.column_config.TextColumn(
                        "Current Fill Rate",
                        help="Fill percentage"
                    ),
                    "Map Trajectory": st.column_config.TextColumn(
                        "Map Trajectory",
                        help="Click here and paste SharePoint trajectory data",
                        max_chars=1000
                    )
                }
            )

            # Update session state immediately when data changes
            if not paste_df.equals(st.session_state.trajectory_paste_df):
                st.session_state.trajectory_paste_df = paste_df.copy()
                # Increment version to force widget refresh on next run
                st.session_state.paste_editor_version += 1
                # Also update the comparison tracker
                st.session_state["_prev_trajectory_paste_df"] = paste_df.copy()
                # Force rerun to refresh the editor widget
                st.rerun()

            # Show helpful message if no data has been pasted yet
            has_data = False
            for _, row in paste_df.iterrows():
                if any(str(cell).strip() != "" for cell in row.values):
                    has_data = True
                    break
            
            if not has_data:
                st.info("💡 **Tip:** If paste (Ctrl+V) doesn't work, try these alternatives:\n"
                       "• Right-click in a cell and select 'Paste'\n"
                       "• Type data manually into the cells\n"
                       "• Copy smaller sections at a time\n"
                       "• Make sure SharePoint data is fully loaded (no 'retrieving data' message)")

            # Buttons on the same row - adjusted width for better text display
            col1, col2, col3 = st.columns([1.0, 0.6, 2.4])
            
            with col1:
                use_button_clicked = st.button("Use Pasted Trajectory Data", key="use_pasted_trajectory", disabled=use_paste_disabled)
            
            with col2:
                clear_button_clicked = st.button("Clear Pasted Data", key="clear_pasted_trajectory")
                
            # Remove green styling and apply custom styling to both buttons
            st.markdown("""
            <style>
            /* Ensure button text doesn't wrap */
            div[data-testid="stButton"] > button {
                white-space: nowrap !important;
                overflow: hidden !important;
                text-overflow: ellipsis !important;
                min-width: fit-content !important;
            }
            </style>
            <script>
            // Apply custom styling to both buttons
            function applyButtonStyling() {
                const buttons = document.querySelectorAll('button');
                buttons.forEach(button => {
                    const buttonText = button.textContent.trim();
                    
                    if (buttonText === 'Use Pasted Trajectory Data') {
                        // Blue styling for Use button
                        button.style.setProperty('background', 'linear-gradient(135deg, #007bff 0%, #0056b3 100%)', 'important');
                        button.style.setProperty('color', 'white', 'important');
                        button.style.setProperty('border', 'none', 'important');
                        button.style.setProperty('font-weight', 'bold', 'important');
                        button.style.setProperty('border-radius', '6px', 'important');
                        button.style.setProperty('box-shadow', '0 3px 6px rgba(0, 123, 255, 0.3)', 'important');
                        button.style.setProperty('transition', 'all 0.2s ease', 'important');
                        
                        button.onmouseenter = function() {
                            this.style.setProperty('background', 'linear-gradient(135deg, #0056b3 0%, #004085 100%)', 'important');
                            this.style.setProperty('transform', 'translateY(-1px)', 'important');
                            this.style.setProperty('box-shadow', '0 4px 8px rgba(0, 123, 255, 0.4)', 'important');
                        };
                        
                        button.onmouseleave = function() {
                            this.style.setProperty('background', 'linear-gradient(135deg, #007bff 0%, #0056b3 100%)', 'important');
                            this.style.setProperty('transform', 'translateY(0)', 'important');
                            this.style.setProperty('box-shadow', '0 3px 6px rgba(0, 123, 255, 0.3)', 'important');
                        };
                    }
                    
                    if (buttonText === 'Clear Pasted Data') {
                        // Red styling for Clear button
                        button.style.setProperty('background', 'linear-gradient(135deg, #dc3545 0%, #c82333 100%)', 'important');
                        button.style.setProperty('color', 'white', 'important');
                        button.style.setProperty('border', 'none', 'important');
                        button.style.setProperty('font-weight', 'bold', 'important');
                        button.style.setProperty('border-radius', '6px', 'important');
                        button.style.setProperty('box-shadow', '0 3px 6px rgba(220, 53, 69, 0.3)', 'important');
                        button.style.setProperty('transition', 'all 0.2s ease', 'important');
                        
                        button.onmouseenter = function() {
                            this.style.setProperty('background', 'linear-gradient(135deg, #c82333 0%, #bd2130 100%)', 'important');
                            this.style.setProperty('transform', 'translateY(-1px)', 'important');
                            this.style.setProperty('box-shadow', '0 4px 8px rgba(220, 53, 69, 0.4)', 'important');
                        };
                        
                        button.onmouseleave = function() {
                            this.style.setProperty('background', 'linear-gradient(135deg, #dc3545 0%, #c82333 100%)', 'important');
                            this.style.setProperty('transform', 'translateY(0)', 'important');
                            this.style.setProperty('box-shadow', '0 3px 6px rgba(220, 53, 69, 0.3)', 'important');
                        };
                    }
                });
            }
            
            // Apply immediately and also after any page updates
            applyButtonStyling();
            setTimeout(applyButtonStyling, 100);
            setTimeout(applyButtonStyling, 500);
            setTimeout(applyButtonStyling, 1000);
            </script>
            """, unsafe_allow_html=True)
            
            if use_button_clicked:
                # Check for required columns before processing
                missing_cols = [col for col in paste_columns if col not in paste_df.columns]
                if missing_cols:
                    st.error(f"The following required columns are missing from the pasted data: {', '.join(missing_cols)}. Please ensure your pasted data includes all required columns exactly as shown.")
                else:
                    # Convert dispatch date to match data format and store in separate key
                    # First, let's see what format the data is actually in by preserving the year
                    dispatch_date_formatted = dispatch_date.strftime("%m-%d-%Y") if dispatch_date else ""
                    st.session_state.dispatch_date_formatted = dispatch_date_formatted
                    # Always reset and update the Planned Driving Trajectories fields
                    st.session_state.trajectories = [
                        {"coords": [], "driver": "", "fill_rate": "", "lbs_to_add": "", "url": "", "added": False}
                        for _ in range(7)
                    ]
                    for i in range(min(7, len(paste_df))):
                        row = paste_df.iloc[i]
                        driver = str(row["Driver/Trailer"]) if pd.notna(row["Driver/Trailer"]) else ""
                        url = str(row["Map Trajectory"]) if pd.notna(row["Map Trajectory"]) else ""
                        fill_rate = str(row["Current Fill Rate"]) if pd.notna(row["Current Fill Rate"]) else ""
                        # Robustly parse numbers, removing commas, spaces, and non-numeric chars
                        def parse_number(val):
                            if pd.isna(val):
                                return 0
                            try:
                                return pd.to_numeric(str(val).replace(",", "").replace(" ", "").strip(), errors="coerce") or 0
                            except Exception:
                                return 0
                        allowed_weight = parse_number(row["Allowed Weight"])
                        gross_weight = parse_number(row["Total Gross Weight"])
                        lbs_to_add = allowed_weight - gross_weight
                        lbs_to_add_fmt = f"{int(round(lbs_to_add)):,}" if not pd.isna(lbs_to_add) else ""
                        st.session_state.trajectories[i] = {
                            "coords": [],
                            "driver": driver,
                            "fill_rate": fill_rate,
                            "lbs_to_add": lbs_to_add_fmt,
                            "url": url,
                            "added": False
                        }
                    
                    # Force widget state updates by incrementing a version counter
                    if "trajectory_widget_version" not in st.session_state:
                        st.session_state.trajectory_widget_version = 0
                    st.session_state.trajectory_widget_version += 1
                    
                    st.success("Pasted trajectory data loaded successfully.")
                    # Do NOT clear or reset the pasted table, just rerun to update UI
                    st.rerun()

            if clear_button_clicked:
                st.session_state.trajectory_paste_df = pd.DataFrame([[""]*len(paste_columns) for _ in range(8)], columns=paste_columns)
                st.session_state["_prev_trajectory_paste_df"] = st.session_state.trajectory_paste_df.copy()
                st.session_state.trajectories = [
                    {"coords": [], "driver": "", "fill_rate": "", "lbs_to_add": "", "url": "", "added": False}
                    for _ in range(7)
                ]
                
                # Force widget state updates by incrementing a version counter
                if "trajectory_widget_version" not in st.session_state:
                    st.session_state.trajectory_widget_version = 0
                st.session_state.trajectory_widget_version += 1
                
                st.rerun()

        st.markdown('<div class="trajectories-label">Planned Driving Trajectories (Google Maps URLs) </div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)  # Add space after the title to prevent overlap
        
        # Initialize widget version if not exists
        if "trajectory_widget_version" not in st.session_state:
            st.session_state.trajectory_widget_version = 0
            
        # Ensure trajectories list has exactly 7 elements
        while len(st.session_state.trajectories) < 7:
            st.session_state.trajectories.append({"coords": [], "driver": "", "fill_rate": "", "lbs_to_add": "", "url": "", "added": False})
        
        trajectory_colors = ['blue', 'green', 'purple', 'orange', 'red', 'cyan', 'magenta']
        
        # Determine how many rows to show based on data availability
        rows_to_show = 7  # Default to all 7 rows
        
        # Check if any data has been populated from paste operation
        has_populated_data = any(
            traj["driver"].strip() or traj["fill_rate"].strip() or traj["lbs_to_add"].strip() or traj["url"].strip()
            for traj in st.session_state.trajectories
        )
        
        if has_populated_data:
            # If data has been populated, only show rows that have at least one field filled
            rows_to_show = 0
            for i in range(7):
                traj = st.session_state.trajectories[i]
                if (traj["driver"].strip() or traj["fill_rate"].strip() or 
                    traj["lbs_to_add"].strip() or traj["url"].strip()):
                    rows_to_show = i + 1
            # Don't add extra empty row - show only rows with data
        
        for i in range(rows_to_show):
            col_driver, col_fill_rate, col_lbs_to_add, col_url, col_button = st.columns([0.15, 0.15, 0.15, 0.45, 0.1])
            with col_driver:
                # Show label only for first row
                if i == 0:
                    st.markdown('<div class="filters-label">Driver Name:</div>', unsafe_allow_html=True)
                driver_name = st.text_input(
                    f"Driver Name {i+1}",
                    value=st.session_state.trajectories[i]["driver"],
                    placeholder=f"Enter driver name {i+1}",
                    label_visibility="hidden",
                    key=f"driver_name_{i}_v{st.session_state.trajectory_widget_version}"
                )
                st.session_state.trajectories[i]["driver"] = driver_name
            with col_fill_rate:
                # Show label only for first row
                if i == 0:
                    st.markdown('<div class="filters-label">Current Fill Rate %:</div>', unsafe_allow_html=True)
                fill_rate = st.text_input(
                    f"Fill Rate {i+1}",
                    value=st.session_state.trajectories[i]["fill_rate"],
                    placeholder="Fill Rate %",
                    label_visibility="hidden",
                    key=f"fill_rate_{i}_v{st.session_state.trajectory_widget_version}"
                )
                st.session_state.trajectories[i]["fill_rate"] = fill_rate
            with col_lbs_to_add:
                # Show label only for first row
                if i == 0:
                    st.markdown('<div class="filters-label">LBS to Add:</div>', unsafe_allow_html=True)
                lbs_to_add = st.text_input(
                    f"LBS to Add {i+1}",
                    value=st.session_state.trajectories[i]["lbs_to_add"],
                    placeholder="LBS to Add",
                    label_visibility="hidden",
                    key=f"lbs_to_add_{i}_v{st.session_state.trajectory_widget_version}"
                )
                st.session_state.trajectories[i]["lbs_to_add"] = lbs_to_add
            with col_url:
                # Show label only for first row
                if i == 0:
                    st.markdown('<div class="filters-label">Google Maps URL:</div>', unsafe_allow_html=True)
                trajectory_url = st.text_input(
                    f"Trajectory URL {i+1}",
                    value=st.session_state.trajectories[i]["url"],
                    placeholder=f"Enter Google Maps directions URL {i+1}",
                    label_visibility="hidden",
                    key=f"trajectory_input_{i}_v{st.session_state.trajectory_widget_version}"
                )
                st.session_state.trajectories[i]["url"] = trajectory_url
            with col_button:
                # Show label only for first row
                if i == 0:
                    st.markdown('<div class="filters-label"> </div>', unsafe_allow_html=True)
                if st.button("Add to Map", key=f"add_trajectory_button_{i}"):
                    # Ensure trajectories list has enough elements
                    while len(st.session_state.trajectories) <= i:
                        st.session_state.trajectories.append({"coords": [], "driver": "", "fill_rate": "", "lbs_to_add": "", "url": "", "added": False})
                    
                    if trajectory_url:
                        addresses = parse_google_maps_url(trajectory_url)
                        if addresses:
                            geolocator = GoogleV3(api_key=st.secrets["GOOGLE_API_KEY"], timeout=10)
                            coords, cache, invalid_addresses = geocode_addresses(addresses, cache, geolocator)
                            if len(coords) >= 2:
                                driving_coords = get_driving_route(addresses, gmaps_client)
                                st.session_state.trajectories[i]["coords"] = driving_coords
                                st.session_state.trajectories[i]["added"] = True
                                if driving_coords:
                                    st.success(f"Trajectory {i+1} added with {len(driving_coords)} points for driver {driver_name or f'Trajectory {i+1}' }.")
                                else:
                                    st.session_state.trajectories[i]["coords"] = coords
                                    st.warning(f"No valid driving route obtained for Trajectory {i+1}. Using geocoded points as fallback.")
                                if invalid_addresses:
                                    st.warning(f"The following addresses could not be geocoded for Trajectory {i+1}:")
                                    for addr in invalid_addresses:
                                        st.write(f"- {addr}")
                            else:
                                st.session_state.trajectories[i]["coords"] = []
                                st.session_state.trajectories[i]["added"] = True
                                st.warning(f"Insufficient valid coordinates for Trajectory {i+1} (need at least 2 points).")
                                if invalid_addresses:
                                    st.warning(f"The following addresses could not be geocoded for Trajectory {i+1}:")
                                    for addr in invalid_addresses:
                                        st.write(f"- {addr}")
                        else:
                            st.session_state.trajectories[i]["coords"] = []
                            st.session_state.trajectories[i]["added"] = True
                            st.warning(f"No valid addresses extracted from the provided Google Maps URL for Trajectory {i+1}.")
                    else:
                        st.session_state.trajectories[i]["coords"] = []
                        st.session_state.trajectories[i]["added"] = True
                        st.info(f"Please enter a Google Maps URL for Trajectory {i+1} to add a driving trajectory.")
                    st.rerun()

        # Align Add All Trajectories button directly under Add Trajectory 7

        col_driver, col_fill_rate, col_lbs_to_add, col_url, col_button = st.columns([0.15, 0.15, 0.15, 0.45, 0.1])
        with col_button:
            add_all = st.button("Add All to Map", key="add_all_trajectories_button", type="primary", help="Add all trajectories at once")
            if add_all:
                # Ensure trajectories list has exactly 7 elements
                while len(st.session_state.trajectories) < 7:
                    st.session_state.trajectories.append({"coords": [], "driver": "", "fill_rate": "", "lbs_to_add": "", "url": "", "added": False})
                
                with st.spinner("Adding all trajectories..."):
                    success_messages = []
                    warning_messages = []
                    for i in range(7):
                        trajectory_url = st.session_state.trajectories[i]["url"]
                        driver_name = st.session_state.trajectories[i]["driver"] or f"Trajectory {i+1}"
                        if trajectory_url:
                            addresses = parse_google_maps_url(trajectory_url)
                            if addresses:
                                geolocator = GoogleV3(api_key=st.secrets["GOOGLE_API_KEY"], timeout=10)
                                coords, cache, invalid_addresses = geocode_addresses(addresses, cache, geolocator)
                                if len(coords) >= 2:
                                    driving_coords = get_driving_route(addresses, gmaps_client)
                                    st.session_state.trajectories[i]["coords"] = driving_coords
                                    st.session_state.trajectories[i]["added"] = True
                                    if driving_coords:
                                        success_messages.append(f"Trajectory {i+1} added with {len(driving_coords)} points for driver {driver_name}.")
                                    else:
                                        st.session_state.trajectories[i]["coords"] = coords
                                        warning_messages.append(f"No valid driving route obtained for Trajectory {i+1}. Using geocoded points as fallback.")
                                    if invalid_addresses:
                                        warning_messages.append(f"The following addresses could not be geocoded for Trajectory {i+1}: {', '.join(invalid_addresses)}")
                                else:
                                    st.session_state.trajectories[i]["coords"] = []
                                    st.session_state.trajectories[i]["added"] = True
                                    warning_messages.append(f"Insufficient valid coordinates for Trajectory {i+1} (need at least 2 points).")
                                    if invalid_addresses:
                                        warning_messages.append(f"The following addresses could not be geocoded for Trajectory {i+1}: {', '.join(invalid_addresses)}")
                            else:
                                st.session_state.trajectories[i]["coords"] = []
                                st.session_state.trajectories[i]["added"] = True
                                warning_messages.append(f"No valid addresses extracted from the provided Google Maps URL for Trajectory {i+1}.")
                        else:
                            st.session_state.trajectories[i]["coords"] = []
                            st.session_state.trajectories[i]["added"] = True
                            warning_messages.append(f"Please enter a Google Maps URL for Trajectory {i+1} to add a driving trajectory.")
                    for msg in success_messages:
                        st.success(msg)
                    for msg in warning_messages:
                        st.warning(msg)
                st.rerun()

    with st.expander("🗺️ Delivery Map", expanded=False):
        # Exclude the selected Dispatch Date and any dates before it from the Map Date Filter options
        # Handle both Paste Table and Upload Excel modes
        dispatch_date_str = None
        dispatch_date_val_paste = st.session_state.get("dispatching_for_paste", None)
        dispatch_date_val_upload = st.session_state.get("dispatching_for", None)
        dispatch_date_formatted = st.session_state.get("dispatch_date_formatted", None)
        
        paste_mode_active = dispatch_date_val_paste is not None or dispatch_date_formatted is not None
        upload_mode_active = dispatch_date_val_upload is not None and not paste_mode_active
        
        if paste_mode_active:
            # Use the pre-formatted date if available, otherwise convert the date input
            if dispatch_date_formatted:
                dispatch_date_str = dispatch_date_formatted
            else:
                try:
                    # Handle both string and datetime objects, convert to MM-DD-YY format
                    if isinstance(dispatch_date_val_paste, str):
                        # Try to parse string date in various formats
                        parsed_date = None
                        for fmt in ["%m-%d-%Y", "%m-%d-%y", "%Y-%m-%d"]:
                            try:
                                parsed_date = pd.to_datetime(dispatch_date_val_paste, format=fmt)
                                break
                            except:
                                continue
                        if parsed_date and not pd.isna(parsed_date):
                            dispatch_date_str = parsed_date.strftime("%m-%d-%Y")
                        else:
                            dispatch_date_str = str(dispatch_date_val_paste).strip().replace(" ", "")
                    else:
                        # Handle datetime object
                        dispatch_date_str = dispatch_date_val_paste.strftime("%m-%d-%Y")
                except Exception:
                    dispatch_date_str = str(dispatch_date_val_paste).strip().replace(" ", "")
        elif upload_mode_active:
            dispatch_date_str = str(dispatch_date_val_upload).strip().replace(" ", "")
        all_dates = sorted(df_geo["Date"].unique(), key=lambda x: pd.to_datetime(x, format="%m-%d-%Y", errors="coerce"))
        valid_dispatch_date = False
        dispatch_date_dt = None
        if (paste_mode_active or upload_mode_active) and dispatch_date_str:
            try:
                dispatch_date_dt = pd.to_datetime(dispatch_date_str, format="%m-%d-%Y", errors="coerce")
                valid_dispatch_date = not pd.isna(dispatch_date_dt)
            except Exception:
                valid_dispatch_date = False
        if valid_dispatch_date:
            # Include the dispatch date and all dates after it (>= instead of >)
            map_date_vals = [d for d in all_dates if pd.to_datetime(d, format="%m-%d-%Y", errors="coerce") >= dispatch_date_dt]
        else:
            map_date_vals = all_dates
        st.markdown('<div class="filters-label">Map Date Filter:</div>', unsafe_allow_html=True)
        # Filter out any default values that are not in map_date_vals
        prev_map_date_filter = st.session_state.map_date_filter if "map_date_filter" in st.session_state else []
        valid_default = [d for d in prev_map_date_filter if d in map_date_vals]
        map_date_choice = st.multiselect(
            "Map Date Filter:",
            map_date_vals,
            default=map_date_vals if not valid_default else valid_default,
            label_visibility="hidden",
            key="map_date_multiselect"
        )
        st.session_state.map_date_filter = map_date_choice

        df_to_plot = df_geo.dropna(subset=["lat_jitter", "lon_jitter"]).copy()
        df_to_plot = df_to_plot[df_to_plot["Date"].isin(map_date_choice)]

        non_plotted_addresses = df_geo[df_geo["lat_jitter"].isna() | df_geo["lon_jitter"].isna()][["Name", "full_address", "Whs", "Date"]].drop_duplicates()

        # Trajectory Visibility Controls
        st.markdown('<div class="filters-label">Trajectory Visibility:</div>', unsafe_allow_html=True)
        st.markdown('<div style="margin-bottom: 10px;"></div>', unsafe_allow_html=True)  # Add spacing
        
        # Get available trajectories with drivers
        available_trajectories = []
        for i, traj in enumerate(st.session_state.trajectories):
            if traj["coords"]:
                driver_name = traj["driver"] or f"Trajectory {i+1}"
                available_trajectories.append(f"{i+1}: {driver_name}")
        
        if available_trajectories:
            # Initialize visible trajectories in session state
            if "visible_trajectories" not in st.session_state:
                st.session_state.visible_trajectories = list(range(len(available_trajectories)))
            
            col1, col2 = st.columns([1, 3])
            
            with col1:
                # All/None buttons on the same line with closer spacing
                btn_col1, btn_col2, btn_col3 = st.columns([1, 1, 2])
                with btn_col1:
                    if st.button("Show All", key="show_all_trajectories"):
                        st.session_state.visible_trajectories = list(range(len(available_trajectories)))
                        st.rerun()
                with btn_col2:
                    if st.button("Hide All", key="hide_all_trajectories"):
                        st.session_state.visible_trajectories = []
                        st.rerun()
                # btn_col3 is empty and acts as spacer
            
            with col2:
                # Individual trajectory checkboxes
                selected_trajectories = []
                cols = st.columns(min(4, len(available_trajectories)))  # Max 4 columns
                
                for i, traj_name in enumerate(available_trajectories):
                    with cols[i % 4]:
                        is_checked = i in st.session_state.visible_trajectories
                        if st.checkbox(traj_name, value=is_checked, key=f"traj_checkbox_{i}"):
                            if i not in selected_trajectories:
                                selected_trajectories.append(i)
                        elif i in st.session_state.visible_trajectories:
                            # Trajectory was unchecked
                            pass
                        else:
                            selected_trajectories.append(i)
                
                # Update visible trajectories based on checkboxes
                new_visible = []
                for i, traj_name in enumerate(available_trajectories):
                    if st.session_state.get(f"traj_checkbox_{i}", False):
                        new_visible.append(i)
                st.session_state.visible_trajectories = new_visible
        
        st.markdown("<br>", unsafe_allow_html=True)  # Add some space

        if not df_to_plot.empty:
            center_lat = df_to_plot["lat_jitter"].mean()
            center_lon = df_to_plot["lon_jitter"].mean()
        else:
            center_lat = WAREHOUSE_COORDS.get(whs, {"lat": 40.7128}).get("lat", 40.7128)
            center_lon = WAREHOUSE_COORDS.get(whs, {"lon": -74.0060}).get("lon", -74.0060)

        m = folium.Map(
            location=[center_lat, center_lon],
            zoom_start=10,
            tiles="OpenStreetMap",
            zoom_control=True
        )

        if whs in WAREHOUSE_COORDS:
            warehouse = WAREHOUSE_COORDS[whs]
            folium.Marker(
                location=[warehouse["lat"], warehouse["lon"]],
                tooltip=folium.Tooltip(f"<b>{warehouse['name']}</b>"),
                icon=folium.Icon(color="red", icon="star", prefix="fa")
            ).add_to(m)

        if not df_to_plot.empty:
            filtered_dates = sorted(df_to_plot["Date"].unique())
            palette = [
                [255, 0, 0], [0, 128, 0], [0, 0, 255], [255, 165, 0], [128, 0, 128],
                [255, 192, 203], [0, 255, 255], [255, 255, 0], [139, 69, 19], [128, 128, 128],
                [255, 99, 71], [0, 191, 255], [75, 0, 130], [173, 255, 47], [220, 20, 60]
            ]
            # Create color mapping with black for the first date
            filtered_cmap = {}
            for i, d in enumerate(filtered_dates):
                if i == 0:
                    # First date gets black color
                    filtered_cmap[d] = [0, 0, 0, 200]  # Black with alpha
                else:
                    # Other dates use the palette (shifted by 1 to skip the red)
                    filtered_cmap[d] = palette[(i-1) % len(palette)] + [200]
            df_to_plot["color"] = df_to_plot["Date"].map(filtered_cmap)

            for row in df_to_plot.to_dict(orient='records'):
                if pd.notna(row["lat_jitter"]) and pd.notna(row["lon_jitter"]):
                    try:
                        tooltip = folium.Tooltip(
                            f"<b>Date:</b> {sanitize_for_html(row['Date'])}<br>"
                            f"<b>Company Name:</b> {sanitize_for_html(row['Name'])}<br>"
                            f"<b>Address:</b> {sanitize_for_html(row['full_address'])}<br>"
                            f"<b>Product:</b> {sanitize_for_html(row['Product'])}<br>"
                            f"<b>Unit:</b> {sanitize_for_html(row['U/M'])}<br>"
                            f"<b>Qty:</b> {row['Qty']}<br>"
                            f"<b>Whs:</b> {row['Whs']}<br>",
                            sticky=True
                        )
                        folium.CircleMarker(
                            location=(row["lat_jitter"], row["lon_jitter"]),
                            radius=8,
                            stroke=False,
                            fill=True,
                            fill_opacity=0.7,
                            fill_color='#%02x%02x%02x' % tuple(row["color"][:3]),
                            tooltip=tooltip
                        ).add_to(m)
                    except Exception:
                        continue

            counts = df_to_plot.groupby("full_address")["Date"].nunique().to_dict()
            for addr, cnt in counts.items():
                if cnt >= 2:
                    coords_found = df_to_plot[df_to_plot["full_address"] == addr][['lat', 'lon']]
                    if not coords_found.empty:
                        lat0, lon0 = coords_found.iloc[0]['lat'], coords_found.iloc[0]['lon']
                        folium.map.Marker(
                            location=(lat0, lon0),
                            icon=DivIcon(
                                icon_size=(30, 30),
                                icon_anchor=(15, 15),
                                html=f'<div style="font-size:12px;font-weight:bold;color:black">{cnt}</div>'
                            )
                        ).add_to(m)

        # Add all trajectories to the map (only visible ones)
        trajectory_colors = ['blue', 'green', 'purple', 'orange', 'red', 'cyan', 'magenta']
        # Enhanced trajectory styling with different patterns and weights
        trajectory_styles = [
            {'color': 'blue', 'weight': 5, 'opacity': 0.9, 'dashArray': None},
            {'color': 'red', 'weight': 5, 'opacity': 0.9, 'dashArray': '10,5'},
            {'color': 'green', 'weight': 6, 'opacity': 0.9, 'dashArray': None},
            {'color': 'purple', 'weight': 5, 'opacity': 0.9, 'dashArray': '15,5,5,5'},
            {'color': 'orange', 'weight': 6, 'opacity': 0.9, 'dashArray': '5,10'},
            {'color': 'darkred', 'weight': 5, 'opacity': 0.9, 'dashArray': None},
            {'color': 'darkgreen', 'weight': 5, 'opacity': 0.9, 'dashArray': '20,5'}
        ]
        
        # Get visible trajectories list
        visible_trajectories = st.session_state.get("visible_trajectories", [])
        
        for i, traj in enumerate(st.session_state.trajectories):
            if len(traj["coords"]) >= 2 and i in visible_trajectories:
                style = trajectory_styles[i % len(trajectory_styles)]
                driver_name = traj['driver'] or f'Trajectory {i+1}'
                
                # Create the polyline with enhanced styling
                polyline = folium.PolyLine(
                    locations=traj["coords"],
                    color=style['color'],
                    weight=style['weight'],
                    opacity=style['opacity'],
                    tooltip=f"📍 {driver_name} | Fill Rate: {traj['fill_rate']} | LBS to Add: {traj['lbs_to_add']}",
                    popup=folium.Popup(
                        f"<b>Driver:</b> {driver_name}<br>"
                        f"<b>Fill Rate:</b> {traj['fill_rate']}<br>"
                        f"<b>LBS to Add:</b> {traj['lbs_to_add']}<br>"
                        f"<b>Trajectory {i+1}</b>",
                        max_width=200
                    )
                )
                
                # Add dash pattern if specified
                if style['dashArray']:
                    polyline.options.update({'dashArray': style['dashArray']})
                
                polyline.add_to(m)
                
                start_coords = traj["coords"][0]
                end_coords = traj["coords"][-1]
                
                folium.CircleMarker(
                    location=end_coords,
                    radius=10,
                    popup=f"WAREHOUSE (END): {driver_name}",
                    tooltip=f"� WAREHOUSE (END): {driver_name}",
                    color='white',
                    weight=2,
                    fill=True,
                    fillColor='red',
                    fillOpacity=0.9
                ).add_to(m)

        all_lats = df_to_plot["lat_jitter"].tolist() + [coord[0] for traj in st.session_state.trajectories for coord in traj["coords"]]
        all_lons = df_to_plot["lon_jitter"].tolist() + [coord[1] for traj in st.session_state.trajectories for coord in traj["coords"]]
        if whs in WAREHOUSE_COORDS:
            all_lats.append(WAREHOUSE_COORDS[whs]["lat"])
            all_lons.append(WAREHOUSE_COORDS[whs]["lon"])
        if all_lats and all_lons:
            min_lat, max_lat = min(all_lats), max(all_lats)
            min_lon, max_lon = min(all_lons), max(all_lons)
            m.fit_bounds([[min_lat, min_lon], [max_lat, max_lon]])
        else:
            if not non_plotted_addresses.empty:
                st.warning("The following addresses could not be plotted due to missing coordinates:")
                for _, row in non_plotted_addresses.iterrows():
                    st.write(f"- {row['Name']}: {row['full_address']} (Whs: {row['Whs']}, Date: {row['Date']})")

        legend_html = """
          <div style="
            position: fixed; top: 10px; left: 50px; width: 200px;
            background: white; padding: 10px; border: 2px solid black;
            z-index: 1000; font-size: 14px; height: auto;
          ">
            <b><strong>Legend:</strong></b><br>
            <b>Dates:</b><br>
        """
        if not df_to_plot.empty:
            for d, col in sorted(filtered_cmap.items()):
                hexc = '#%02x%02x%02x' % tuple(col[:3])
                day_name = pd.to_datetime(d, format="%m-%d-%Y").strftime("%a")
                legend_html += (
                    f'<i style="background:{hexc};width:12px;height:12px;display:inline-block;margin-right:2px;"></i> {sanitize_for_html(d)} ({day_name})<br>'
                )
        if whs in WAREHOUSE_COORDS:
            legend_html += (
                '<i style="color:red;font-family:FontAwesome;font-size:12px;margin-right:2px;"></i> '
                f'{sanitize_for_html(WAREHOUSE_COORDS[whs]["name"])}<br>'
            )
        legend_html += "<b>Trajectories:</b><br>"
        trajectory_legend_styles = [
            {'color': 'blue', 'pattern': '━━━'},
            {'color': 'red', 'pattern': '━ ━'},
            {'color': 'green', 'pattern': '━━━'},
            {'color': 'purple', 'pattern': '━ ━ ━'},
            {'color': 'orange', 'pattern': '━ ━━'},
            {'color': 'darkred', 'pattern': '━━━'},
            {'color': 'darkgreen', 'pattern': '━━ ━'}
        ]
        
        # Get visible trajectories list
        visible_trajectories = st.session_state.get("visible_trajectories", [])
        
        for i, traj in enumerate(st.session_state.trajectories):
            if traj["coords"] and i in visible_trajectories:
                style = trajectory_legend_styles[i % len(trajectory_legend_styles)]
                driver_name = traj["driver"] or f"Trajectory {i+1}"
                legend_html += (
                    f'<div style="margin:1px 0;white-space:nowrap;">'
                    f'<span style="background:{style["color"]};color:white;padding:1px 4px;border-radius:3px;font-size:10px;font-weight:bold;margin-right:3px;">{i+1}</span>'
                    f'<span style="color:{style["color"]};font-weight:bold;font-family:monospace;font-size:12px;margin-right:5px;">{style["pattern"]}</span>'
                    f'<span style="font-size:12px;">{sanitize_for_html(driver_name)}</span>'
                    f'</div>'
                )
        legend_html += "</div>"
        m.get_root().html.add_child(folium.Element(legend_html))

        components.html(m._repr_html_(), height=800)

    with st.expander("🚛 Nearby Deliveries by Driver", expanded=False):
        st.markdown('<div class="filters-label">Proximity Threshold (miles)</div>', unsafe_allow_html=True)
        delivery_map_threshold_miles = st.slider(
            "Proximity Threshold (miles)",
            min_value=1, max_value=150, value=st.session_state.delivery_map_threshold_miles,
            label_visibility="hidden",
            key="delivery_map_threshold_slider"
        )
        st.session_state.delivery_map_threshold_miles = delivery_map_threshold_miles

        st.markdown('<div class="filters-label">Filter Candidates:</div>', unsafe_allow_html=True)
        st.markdown('<div style="font-size: 12px; font-style: italic;">Showing Real Candidates only (Calculated > 0)</div>', unsafe_allow_html=True)
        st.markdown('<div style="font-size: 12px; font-style: italic;">Note: Calculated = On-Hand - (Allocated Qty + Open Orders + C Qty)</div>', unsafe_allow_html=True)
        # --- Download All and Email button at top ---
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Open email and generate Download file for all Drivers", key="download_email_all_drivers"):
            wb = Workbook()
            wb.remove(wb.active)
            file_has_data = False
            invalid_chars = r'[\\/*?:\[\]]'
            driver_lines = []
            
            # Create filtered date list excluding dispatch date for nearby deliveries
            dispatch_date = getattr(st.session_state, 'dispatching_for', None)
            filtered_dates_for_nearby = st.session_state.map_date_filter.copy()
            
            if dispatch_date:
                # Remove dispatch date from the filter for nearby deliveries
                filtered_dates_for_nearby = [date for date in filtered_dates_for_nearby if date != dispatch_date]
            
            for i, traj in enumerate(st.session_state.trajectories):
                driver_name = traj["driver"] or f"Trajectory {i+1}"
                safe_title = re.sub(invalid_chars, '_', driver_name)[:31]
                
                # Get all nearby deliveries first
                nearby_deliveries_df = find_nearby_deliveries(
                    st.session_state.df_geo,
                    traj["coords"],
                    st.session_state.delivery_map_threshold_miles,
                    show_real_only=True,
                    map_date_filter=st.session_state.map_date_filter
                )
                
                # Filter out the dispatch date from the results
                dispatch_date_obj = st.session_state.get('dispatching_for_paste', None)
                dispatch_date_legacy = getattr(st.session_state, 'dispatching_for', None)
                
                # Convert date object to string format for comparison
                dispatch_date_str = None
                if dispatch_date_obj:
                    dispatch_date_str = dispatch_date_obj.strftime("%m-%d-%Y")
                elif dispatch_date_legacy:
                    dispatch_date_str = dispatch_date_legacy
                    
                if dispatch_date_str and not nearby_deliveries_df.empty and 'Date' in nearby_deliveries_df.columns:
                    # Filter out the dispatch date
                    nearby_deliveries_df = nearby_deliveries_df[
                        nearby_deliveries_df['Date'] != dispatch_date_str
                    ].copy()
                if not nearby_deliveries_df.empty:
                    file_has_data = True
                    ws = wb.create_sheet(title=safe_title)
                    # Compose row 1: Each value in its own column, avoid duplicate 'Dispatching for:'
                    dispatching_for = getattr(st.session_state, 'dispatching_for', None)
                    fill_rate = traj["fill_rate"] or "N/A"
                    lbs_to_add = traj["lbs_to_add"] or "N/A"
                    # Clean up dispatching_for value
                    clean_dispatching_for = dispatching_for
                    if dispatching_for and dispatching_for.strip().lower().startswith("dispatching for:"):
                        clean_dispatching_for = dispatching_for.strip()[len("Dispatching for:"):].strip()
                    # Write summary in A1, B1, C1
                    ws.cell(row=1, column=1, value=f"Dispatching for: {clean_dispatching_for}" if clean_dispatching_for else "")
                    ws.cell(row=1, column=2, value=f"Current Fill Rate: {fill_rate}")
                    ws.cell(row=1, column=3, value=f"LBS to Add: {lbs_to_add}")
                    # Leave row 2 blank for spacing
                    # Write headers in row 3, left-aligned
                    for col_idx, header in enumerate(nearby_deliveries_df.columns, start=1):
                        cell = ws.cell(row=3, column=col_idx, value=header)
                        cell.alignment = Alignment(horizontal='left')
                    # Write data starting from row 4
                    for row_idx, row in enumerate(nearby_deliveries_df.values, start=4):
                        for col_idx, value in enumerate(row, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column].width = adjusted_width
                    # Compose driver line for email
                    # Try to split driver name into name and truck/trailer if possible
                    parts = driver_name.split()
                    if len(parts) >= 2 and "/" in parts[-1]:
                        driver_base = " ".join(parts[:-1])
                        truck_trailer = parts[-1]
                        driver_line = f"Driver: {driver_base} {truck_trailer} | Fill Rate: {fill_rate} | LBS to Add: {lbs_to_add}"
                    else:
                        driver_line = f"Driver: {driver_name} | Fill Rate: {fill_rate} | LBS to Add: {lbs_to_add}"
                    driver_lines.append(driver_line)
            if file_has_data:
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                file_name = f"{whs}_All_Nearby_Deliveries_{current_date}_{current_time}.xlsx"
                # Compose subject with dispatch date - check multiple sources
                dispatching_for = getattr(st.session_state, 'dispatching_for', None)
                dispatching_for_paste = getattr(st.session_state, 'dispatching_for_paste', None)
                
                # Get dispatch date from any available source
                dispatch_date_display = None
                if dispatching_for_paste:
                    # Convert date object to string if needed
                    if hasattr(dispatching_for_paste, 'strftime'):
                        dispatch_date_display = dispatching_for_paste.strftime("%m-%d-%Y")
                    else:
                        dispatch_date_display = str(dispatching_for_paste)
                elif dispatching_for:
                    dispatch_date_display = dispatching_for
                
                # Always include dispatch date in subject if available
                if dispatch_date_display:
                    subject = f"{whs} | All Nearby Delivery Candidates | Dispatch Date: {dispatch_date_display}"
                else:
                    subject = f"{whs} | All Nearby Delivery Candidates"
                # Compose email body in requested format
                body_lines = [
                    "Please find all nearby delivery candidates for all drivers based on planned trajectories attached.",
                    "------------------------------"
                ]
                if dispatch_date_display:
                    # Use the unified dispatch date display
                    clean_dispatch_date = dispatch_date_display
                    if clean_dispatch_date.strip().lower().startswith("dispatching for:"):
                        clean_dispatch_date = clean_dispatch_date.strip()[len("Dispatching for:"):].strip()
                    body_lines.append(f"Dispatching for: {clean_dispatch_date}")
                    body_lines.append("------------------------------")
                if driver_lines:
                    body_lines.extend(driver_lines)
                body = "\n".join(body_lines)
                encoded_subject = urllib.parse.quote(subject)
                encoded_body = urllib.parse.quote(body)
                warehouse_emails = {
                    "D01": "nolwoodcsdetroit@pvschemicals.com",
                    "S01": "nolwoodcscincinnati@pvschemicals.com",
                    "C01": "nolwoodcscleveland@pvschemicals.com",
                    "P01": "nolwoodcspittsburgh@pvschemicals.com"
                }
                to_email = warehouse_emails.get(whs, "")
                mailto_link = f"mailto:{to_email}?subject={encoded_subject}&body={encoded_body}"
                # Open email first
                components.html(
                    f"""
                    <script>
                        window.location.href = '{mailto_link}';
                    </script>
                    """,
                    height=0
                )
                # Then download file
                st.download_button(
                    label="Download All Nearby Deliveries Excel File",
                    data=output.getvalue(),
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_all_nearby_deliveries_file"
                )
            else:
                st.warning("No nearby deliveries found for any driver.")

        # Show nearby deliveries only for drivers that have data
        for i, traj in enumerate(st.session_state.trajectories):
            driver_name = traj["driver"]
            
            # Skip if no driver name is provided
            if not driver_name or driver_name.strip() == "":
                continue
                
            # Get all nearby deliveries first
            nearby_deliveries_df = find_nearby_deliveries(
                st.session_state.df_geo,
                traj["coords"],
                st.session_state.delivery_map_threshold_miles,
                show_real_only=True,
                map_date_filter=st.session_state.map_date_filter
            )
            
            # Filter out the dispatch date from the results
            dispatch_date_obj = st.session_state.get('dispatching_for_paste', None)
            dispatch_date_legacy = getattr(st.session_state, 'dispatching_for', None)
            
            # Convert date object to string format for comparison
            dispatch_date_str = None
            if dispatch_date_obj:
                dispatch_date_str = dispatch_date_obj.strftime("%m-%d-%Y")
            elif dispatch_date_legacy:
                dispatch_date_str = dispatch_date_legacy
                
            if dispatch_date_str and not nearby_deliveries_df.empty and 'Date' in nearby_deliveries_df.columns:
                # Simple direct comparison
                nearby_deliveries_df = nearby_deliveries_df[
                    nearby_deliveries_df['Date'] != dispatch_date_str
                ].copy()
            fill_rate = traj["fill_rate"] or "N/A"
            # Format LBS to add with thousand separator, no decimals
            try:
                lbs_to_add_val = float(traj["lbs_to_add"].replace(",", "")) if isinstance(traj["lbs_to_add"], str) else float(traj["lbs_to_add"])
                lbs_to_add_fmt = f"{int(round(lbs_to_add_val)):,}"
            except (ValueError, TypeError, AttributeError):
                lbs_to_add_fmt = "N/A"
            # Get dispatching_for for this section
            dispatching_for = getattr(st.session_state, 'dispatching_for', None)
            # Avoid duplicate 'Dispatching for:'
            if dispatching_for:
                clean_dispatching_for = dispatching_for
                if clean_dispatching_for.strip().lower().startswith("dispatching for:"):
                    clean_dispatching_for = clean_dispatching_for.strip()[len("Dispatching for:"):].strip()
                dispatch_line = f"Dispatching for: {clean_dispatching_for} | Current Fill Rate: {fill_rate} | LBS to Add: {lbs_to_add_fmt}"
            else:
                dispatch_line = f"Current Fill Rate: {fill_rate} | LBS to Add: {lbs_to_add_fmt}"
            st.markdown(f"<h4>Nearby Deliveries for {driver_name}</h4>", unsafe_allow_html=True)
            st.markdown(f"<b>{dispatch_line}</b>", unsafe_allow_html=True)
            if not nearby_deliveries_df.empty:
                st.dataframe(
                    nearby_deliveries_df, 
                    width='stretch', 
                    height=400, 
                    hide_index=True,
                    column_config={
                        "Distance": st.column_config.NumberColumn(
                            "Distance",
                            help="Distance from/to Trajectory in miles"
                        ),
                        "Calculated": st.column_config.NumberColumn(
                            "Calculated",
                            help="On-Hand - (Allocated Qty + Open Orders + C Qty)"
                        ),
                        "Open Orders": st.column_config.NumberColumn(
                            "Open Orders",
                            help="Future item orders that will take place before listed order"
                        )
                    }
                )
            else:
                st.info(f"No deliveries found within {st.session_state.delivery_map_threshold_miles} miles of the planned driving trajectory for {driver_name}.")

    with st.expander("📦 Overdue Containers Recovery", expanded=False):
        overdue_recovery_df = pd.DataFrame()

        if not st.session_state.overdue_containers_df.empty:
            overdue_df = st.session_state.overdue_containers_df.copy()
            overdue_df["Bill to Name"] = overdue_df["Bill to Name"].astype(str).str.strip()
            overdue_df["Warehouse"] = overdue_df["Warehouse"].astype(str).str.strip()

            dispatch_filtered = df[df["Whs"] == whs].copy()
            dispatch_filtered["Name"] = dispatch_filtered["Name"].astype(str).str.strip()
            dispatch_filtered["full_address"] = (
                dispatch_filtered["Address 1"].fillna('').str.strip() + ", " +
                dispatch_filtered["City"].fillna('').str.strip() + ", " +
                dispatch_filtered["Sta"].fillna('').str.strip()
            )

            overdue_recovery = []

            for _, dispatch_row in dispatch_filtered.iterrows():
                name = dispatch_row["Name"]
                matching_overdue = overdue_df[
                    (overdue_df["Bill to Name"] == name) &
                    (overdue_df["Warehouse"] == whs)
                ]
                if not matching_overdue.empty:
                    try:
                        grouped = matching_overdue.groupby("Bill to Name")[["Tote Number", "Last Sale Date"]].apply(
                            lambda g: "; ".join(
                                f"{tote} ({f'{date.zfill(6)[:2]}-{date.zfill(6)[2:4]}-{date.zfill(6)[4:6]}' if date.replace('0', '').isdigit() and len(date) <= 6 else 'Invalid'})"
                                for tote, date in zip(g["Tote Number"].astype(str), g["Last Sale Date"].astype(str))
                            )
                        ).reset_index(name="Tote and Last Sale Date")

                        co_numbers = matching_overdue.groupby("Bill to Name")["CO Number"].apply(
                            lambda x: ";".join(x.astype(str).unique())
                        ).reset_index(name="CO Number")

                        merged = pd.merge(grouped, co_numbers, on="Bill to Name")

                        for _, group_row in merged.iterrows():
                            overdue_recovery.append({
                                "Dispatch Date": dispatch_row["Date"],
                                "CO no": dispatch_row["CO no"],
                                "Name": dispatch_row["Name"],
                                "Address": dispatch_row["full_address"],
                                "Resp": dispatch_row.get("Resp", ""),
                                "Salespers": dispatch_row.get("Salespers", ""),
                                "Tote and Last Sale Date": group_row["Tote and Last Sale Date"],
                                "CO Number": group_row["CO Number"]
                            })
                    except Exception as e:
                        st.warning(f"Error processing overdue containers: {e}")

            overdue_recovery_df = pd.DataFrame(overdue_recovery)

           

            if not overdue_recovery_df.empty:
                tote_numbers = set()
                for tote_date_str in overdue_recovery_df["Tote and Last Sale Date"].dropna():
                    totes = [t.split(" (", 1)[0] for t in tote_date_str.split("; ")]
                    tote_numbers.update(totes)
                tote_count = len(tote_numbers)

                overdue_recovery_df = overdue_recovery_df.drop_duplicates(subset=["Dispatch Date", "Name", "Address"])

                overdue_recovery_df["Dispatch Date"] = pd.to_datetime(overdue_recovery_df["Dispatch Date"], format="%m-%d-%Y", errors="coerce")
                overdue_recovery_df = overdue_recovery_df.sort_values("Dispatch Date")
                overdue_recovery_df["Dispatch Date"] = overdue_recovery_df["Dispatch Date"].dt.strftime("%m-%d-%Y")

                st.write(f"{tote_count} overdue containers can be recovered based on future dispatch activities.")
                st.dataframe(
                    overdue_recovery_df,
                    width='stretch',
                    height=400,
                    hide_index=True,
                    column_config={
                        "Dispatch Date": st.column_config.TextColumn("Dispatch Date"),
                        "CO no": st.column_config.TextColumn("CO no"),
                        "Name": st.column_config.TextColumn("Name"),
                        "Address": st.column_config.TextColumn("Address"),
                        "Resp": st.column_config.TextColumn("Resp"),
                        "Salespers": st.column_config.TextColumn("Salespers"),
                        "Tote and Last Sale Date": st.column_config.TextColumn("Tote and Last Sale Date"),
                        "CO Number": st.column_config.TextColumn("CO Number")
                    }
                )

                excel_content = generate_excel_file(overdue_recovery_df, sheet_name="Overdue Containers", align_left=True)
                if excel_content:
                    subject = f"{whs} | Overdue Containers Recovery Report {current_date}"
                    body = "Please find the overdue containers recovery report attached.\nThe overdue containers recovery report is attached. In preparation for tomorrow's deliveries, inform the customers about the upcoming delivery and ensure the empty totes are ready for the drivers' pickup.\nThank you."
                    encoded_subject = urllib.parse.quote(subject)
                    encoded_body = urllib.parse.quote(body)

                    to_recipients = list(set(transform_username(recipient) + "@pvschemicals.com" for recipient in overdue_recovery_df["Resp"].dropna().str.strip() if recipient))
                    cc_recipients = list(set(transform_username(recipient) + "@pvschemicals.com" for recipient in overdue_recovery_df["Salespers"].dropna().str.strip() if recipient))
                    to_recipients_str = ";".join(to_recipients[:10])
                    cc_recipients_str = ";".join(cc_recipients[:10])

                    mailto_link = f"mailto:{to_recipients_str}?cc={cc_recipients_str}&subject={encoded_subject}&body={encoded_body}"

                    file_name = f"{whs}_Overdue_Containers_Recovery_{current_date}_{current_time}.xlsx"

                    if st.download_button(
                        label="Open Outlook and Download File",
                        data=excel_content,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="overdue_email_button"
                    ):
                        components.html(
                            f"""
                            <script>
                                window.location.href = "{mailto_link}";
                            </script>
                            """,
                            height=0
                        )
                        st.session_state.overdue_email_triggered = False
                else:
                    st.warning("No data available to download for Overdue Containers Recovery.")
            else:
                st.info("No overdue containers found for recovery based on current filters.")

# End of Nearby Deliveries Planning section